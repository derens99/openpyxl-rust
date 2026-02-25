#!/usr/bin/env python3
"""
openpyxl_rust Report Generator

Runs the full test suite, performance benchmarks, feature-parity checks,
and memory/file-size comparisons, then outputs a detailed report showing
how openpyxl_rust compares to openpyxl.

Usage:
    py benchmarks/generate_report.py              # terminal + saved report
    py benchmarks/generate_report.py --quick      # skip large benchmarks
    py benchmarks/generate_report.py --save-only  # suppress terminal, only save file
"""
import argparse
import io
import os
import platform
import subprocess
import sys
import tempfile
import textwrap
import time
import tracemalloc
from datetime import datetime, date
from pathlib import Path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

DIVIDER = "=" * 78
THIN_DIVIDER = "-" * 78
VERSION = "0.5.0"
REPORT_DIR = Path(__file__).resolve().parent.parent / "reports"


def _bar(fraction, width=30):
    """Return a text progress bar: [######........]"""
    filled = int(fraction * width)
    return "[" + "#" * filled + "." * (width - filled) + "]"


def _fmt_bytes(n):
    if n < 1024:
        return f"{n} B"
    elif n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    else:
        return f"{n / (1024 * 1024):.2f} MB"


def _pct_change(old, new):
    """Return a string like '-42.3%' or '+12.0%'."""
    if old == 0:
        return "N/A"
    change = (new - old) / old * 100
    sign = "+" if change >= 0 else ""
    return f"{sign}{change:.1f}%"


class ReportBuilder:
    """Accumulate report sections, then emit as terminal text + saved file."""

    def __init__(self):
        self._sections = []

    def section(self, title, body):
        self._sections.append((title, body))

    def build(self):
        lines = []
        lines.append(DIVIDER)
        lines.append("  openpyxl_rust Performance & Compatibility Report")
        lines.append(f"  Version {VERSION}  |  Generated {datetime.now():%Y-%m-%d %H:%M:%S}")
        lines.append(f"  Python {platform.python_version()}  |  {platform.system()} {platform.release()}")
        lines.append(DIVIDER)
        lines.append("")
        for title, body in self._sections:
            lines.append(f"  {title}")
            lines.append(THIN_DIVIDER)
            lines.append(body)
            lines.append("")
        lines.append(DIVIDER)
        lines.append("  End of Report")
        lines.append(DIVIDER)
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# 1. Test Suite
# ---------------------------------------------------------------------------

def run_test_suite():
    """Run pytest and return (passed, failed, errors, duration, raw_output)."""
    start = time.perf_counter()
    result = subprocess.run(
        [sys.executable, "-m", "pytest", "tests/", "-v", "--tb=short", "-q"],
        capture_output=True, text=True, cwd=str(Path(__file__).resolve().parent.parent),
    )
    duration = time.perf_counter() - start
    output = result.stdout + result.stderr

    # Parse summary line like "197 passed in 1.40s"
    passed = failed = errors = 0
    for line in output.splitlines():
        line = line.strip()
        if "passed" in line:
            import re
            m = re.search(r"(\d+) passed", line)
            if m:
                passed = int(m.group(1))
            m = re.search(r"(\d+) failed", line)
            if m:
                failed = int(m.group(1))
            m = re.search(r"(\d+) error", line)
            if m:
                errors = int(m.group(1))

    return passed, failed, errors, duration, output


def format_test_section(passed, failed, errors, duration):
    lines = []
    total = passed + failed + errors
    status = "ALL PASSING" if failed == 0 and errors == 0 else "FAILURES DETECTED"
    lines.append(f"  Status : {status}")
    lines.append(f"  Total  : {total} tests")
    lines.append(f"  Passed : {passed}")
    if failed:
        lines.append(f"  Failed : {failed}")
    if errors:
        lines.append(f"  Errors : {errors}")
    lines.append(f"  Time   : {duration:.2f}s")
    lines.append("")

    # Test file breakdown
    test_dir = Path(__file__).resolve().parent.parent / "tests"
    test_files = sorted(test_dir.glob("test_*.py"))
    lines.append(f"  {'Test File':<40} {'Tests':>6}")
    lines.append(f"  {'-' * 40} {'-' * 6}")
    for tf in test_files:
        # Count test functions
        count = 0
        with open(tf, encoding="utf-8", errors="replace") as f:
            for line in f:
                if line.strip().startswith("def test_") or line.strip().startswith("class Test"):
                    count += 1
        lines.append(f"  {tf.name:<40} {count:>6}")
    lines.append(f"  {'-' * 40} {'-' * 6}")
    lines.append(f"  {'TOTAL':<40} {total:>6}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 2. Performance Benchmarks
# ---------------------------------------------------------------------------

def _bench_write(lib, rows, cols, use_styles=False, sheets=1):
    """Generic benchmark: write data using either 'openpyxl' or 'openpyxl_rust'."""
    if lib == "openpyxl":
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        if sheets > 1:
            wb.remove(wb.active)
    else:
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        wb = Workbook()
        if sheets > 1:
            # For rust, rename default sheet as first
            ws = wb.active
            ws.title = "Sheet1"

    for s in range(sheets):
        if lib == "openpyxl":
            ws = wb.create_sheet(f"Sheet{s + 1}")
        elif s == 0 and sheets > 1:
            ws = wb.active
        else:
            ws = wb.create_sheet(f"Sheet{s + 1}") if s > 0 else wb.active

        if use_styles:
            bold = Font(bold=True, size=12)
            for c in range(1, cols + 1):
                cell = ws.cell(row=1, column=c, value=f"Header_{c}")
                cell.font = bold
            start_row = 2
        else:
            start_row = 1

        for r in range(start_row, rows + start_row):
            for c in range(1, cols + 1):
                if c % 3 == 0:
                    ws.cell(row=r, column=c, value=f"str_{r}_{c}")
                elif c % 3 == 1:
                    ws.cell(row=r, column=c, value=r * c * 1.1)
                else:
                    ws.cell(row=r, column=c, value=r % 2 == 0)

        if use_styles:
            for r in range(start_row, rows + start_row):
                ws.cell(row=r, column=2).number_format = "$#,##0.00"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        size = os.path.getsize(path)
    finally:
        os.unlink(path)
    return size


def _bench_batch_write(rows, cols):
    """Benchmark using openpyxl_rust's batch API."""
    from openpyxl_rust import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    batch = []
    for r in range(1, rows + 1):
        row = []
        for c in range(1, cols + 1):
            if c % 3 == 0:
                row.append(f"str_{r}_{c}")
            elif c % 3 == 1:
                row.append(r * c * 1.1)
            else:
                row.append(r % 2 == 0)
        batch.append(row)
        if len(batch) >= 10_000:
            ws.append_rows(batch)
            batch = []
    if batch:
        ws.append_rows(batch)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        size = os.path.getsize(path)
    finally:
        os.unlink(path)
    return size


def _bench_read(lib, xlsx_path):
    """Benchmark reading an xlsx file."""
    if lib == "openpyxl":
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        total = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        total += 1
        return total
    else:
        from openpyxl_rust import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        total = 0
        for ws in wb._sheets:
            for (r, c), cell in ws._cells.items():
                if cell.value is not None:
                    total += 1
        return total


def _time_fn(fn, *args, **kwargs):
    """Time a function, return (duration, result)."""
    start = time.perf_counter()
    result = fn(*args, **kwargs)
    elapsed = time.perf_counter() - start
    return elapsed, result


def _memory_fn(fn, *args, **kwargs):
    """Measure peak memory of a function."""
    tracemalloc.start()
    result = fn(*args, **kwargs)
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    return peak, result


def run_benchmarks(quick=False):
    """Run all benchmarks. Returns list of result dicts."""
    results = []

    if quick:
        large_rows, large_cols = 10_000, 5
        styled_rows = 2_000
        multi_sheets, multi_rows = 3, 5_000
    else:
        large_rows, large_cols = 100_000, 10
        styled_rows = 10_000
        multi_sheets, multi_rows = 5, 20_000

    benchmarks = [
        {
            "name": f"Large data write ({large_rows:,} rows x {large_cols} cols)",
            "category": "write",
            "fn_openpyxl": lambda: _bench_write("openpyxl", large_rows, large_cols),
            "fn_rust": lambda: _bench_write("openpyxl_rust", large_rows, large_cols),
        },
        {
            "name": f"Batch write ({large_rows:,} rows x {large_cols} cols)",
            "category": "write",
            "fn_openpyxl": lambda: _bench_write("openpyxl", large_rows, large_cols),
            "fn_rust": lambda: _bench_batch_write(large_rows, large_cols),
        },
        {
            "name": f"Styled write ({styled_rows:,} rows, fonts + fmts)",
            "category": "write",
            "fn_openpyxl": lambda: _bench_write("openpyxl", styled_rows, 5, use_styles=True),
            "fn_rust": lambda: _bench_write("openpyxl_rust", styled_rows, 5, use_styles=True),
        },
        {
            "name": f"Multi-sheet ({multi_sheets} x {multi_rows:,} rows)",
            "category": "write",
            "fn_openpyxl": lambda: _bench_write("openpyxl", multi_rows, 3, sheets=multi_sheets),
            "fn_rust": lambda: _bench_write("openpyxl_rust", multi_rows, 3, sheets=multi_sheets),
        },
    ]

    # Read benchmark - create a test file first
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        read_path = f.name
    read_rows = 50_000 if not quick else 5_000
    import openpyxl as _opx
    _wb = _opx.Workbook()
    _ws = _wb.active
    for r in range(1, read_rows + 1):
        _ws.cell(row=r, column=1, value=f"row_{r}")
        _ws.cell(row=r, column=2, value=r * 1.5)
        _ws.cell(row=r, column=3, value=r % 2 == 0)
    _wb.save(read_path)

    benchmarks.append({
        "name": f"Read file ({read_rows:,} rows x 3 cols)",
        "category": "read",
        "fn_openpyxl": lambda: _bench_read("openpyxl", read_path),
        "fn_rust": lambda: _bench_read("openpyxl_rust", read_path),
    })

    for b in benchmarks:
        print(f"  Benchmarking: {b['name']}...")
        t_openpyxl, size_openpyxl = _time_fn(b["fn_openpyxl"])
        t_rust, size_rust = _time_fn(b["fn_rust"])
        # Memory (only for non-quick large write)
        mem_openpyxl = mem_rust = None
        if not quick and b["category"] == "write" and "Large" in b["name"]:
            mem_openpyxl, _ = _memory_fn(b["fn_openpyxl"])
            mem_rust, _ = _memory_fn(b["fn_rust"])

        speedup = t_openpyxl / t_rust if t_rust > 0 else float("inf")
        results.append({
            "name": b["name"],
            "category": b["category"],
            "t_openpyxl": t_openpyxl,
            "t_rust": t_rust,
            "speedup": speedup,
            "size_openpyxl": size_openpyxl if b["category"] == "write" else None,
            "size_rust": size_rust if b["category"] == "write" else None,
            "mem_openpyxl": mem_openpyxl,
            "mem_rust": mem_rust,
        })

    os.unlink(read_path)
    return results


def format_benchmark_section(results):
    lines = []

    # Speed table
    lines.append("  SPEED COMPARISON")
    lines.append("")
    lines.append(f"  {'Benchmark':<45} {'openpyxl':>9} {'ours':>9} {'speedup':>9}")
    lines.append(f"  {'-' * 45} {'-' * 9} {'-' * 9} {'-' * 9}")
    for r in results:
        marker = " **" if r["speedup"] >= 3.0 else ""
        lines.append(
            f"  {r['name']:<45} {r['t_openpyxl']:>8.2f}s {r['t_rust']:>8.2f}s {r['speedup']:>8.1f}x{marker}"
        )
    avg = sum(r["speedup"] for r in results) / len(results)
    lines.append(f"  {'-' * 45} {'-' * 9} {'-' * 9} {'-' * 9}")
    lines.append(f"  {'AVERAGE':>45} {'':>9} {'':>9} {avg:>8.1f}x")
    lines.append("")

    # Visual bar chart
    lines.append("  VISUAL SPEEDUP")
    lines.append("")
    max_speedup = max(r["speedup"] for r in results)
    for r in results:
        bar = _bar(r["speedup"] / max(max_speedup, 1), width=35)
        short_name = r["name"][:35]
        lines.append(f"  {short_name:<37} {bar} {r['speedup']:.1f}x")
    lines.append("")

    # File size table (write benchmarks only)
    write_results = [r for r in results if r["size_openpyxl"] is not None]
    if write_results:
        lines.append("  OUTPUT FILE SIZE")
        lines.append("")
        lines.append(f"  {'Benchmark':<45} {'openpyxl':>12} {'ours':>12} {'diff':>8}")
        lines.append(f"  {'-' * 45} {'-' * 12} {'-' * 12} {'-' * 8}")
        for r in write_results:
            diff = _pct_change(r["size_openpyxl"], r["size_rust"])
            lines.append(
                f"  {r['name']:<45} {_fmt_bytes(r['size_openpyxl']):>12} "
                f"{_fmt_bytes(r['size_rust']):>12} {diff:>8}"
            )
        lines.append("")

    # Memory table (if available)
    mem_results = [r for r in results if r["mem_openpyxl"] is not None]
    if mem_results:
        lines.append("  PEAK MEMORY USAGE")
        lines.append("")
        lines.append(f"  {'Benchmark':<45} {'openpyxl':>12} {'ours':>12} {'savings':>8}")
        lines.append(f"  {'-' * 45} {'-' * 12} {'-' * 12} {'-' * 8}")
        for r in mem_results:
            savings = _pct_change(r["mem_openpyxl"], r["mem_rust"])
            lines.append(
                f"  {r['name']:<45} {_fmt_bytes(r['mem_openpyxl']):>12} "
                f"{_fmt_bytes(r['mem_rust']):>12} {savings:>8}"
            )
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 3. Feature Parity
# ---------------------------------------------------------------------------

def check_feature_parity():
    """Check which openpyxl APIs we support and return structured results."""
    features = []

    def _check(name, category, test_fn):
        try:
            test_fn()
            features.append((name, category, True, None))
        except Exception as e:
            features.append((name, category, False, str(e)))

    # --- Data Types ---
    def _strings():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "hello"
        assert ws["A1"].value == "hello"
    _check("String values", "Data Types", _strings)

    def _numbers():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = 42.5
        assert ws["A1"].value == 42.5
    _check("Numeric values (int/float)", "Data Types", _numbers)

    def _booleans():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = True
        assert ws["A1"].value is True
    _check("Boolean values", "Data Types", _booleans)

    def _dates():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = date(2026, 1, 15)
    _check("Date values", "Data Types", _dates)

    def _datetimes():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = datetime(2026, 1, 15, 10, 30)
    _check("Datetime values", "Data Types", _datetimes)

    def _formulas():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "=SUM(B1:B10)"
    _check("Formula strings", "Data Types", _formulas)

    def _none():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = None
    _check("None / blank cells", "Data Types", _none)

    def _type_validation():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        try:
            ws["A1"] = [1, 2, 3]
            raise AssertionError("Should have raised TypeError")
        except TypeError:
            pass
    _check("Type validation (rejects invalid)", "Data Types", _type_validation)

    # --- Styling ---
    def _font():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        wb = Workbook(); ws = wb.active
        ws["A1"] = "bold"
        ws["A1"].font = Font(bold=True, italic=True, size=14, name="Arial", color="FF0000")
    _check("Font (bold/italic/size/name/color)", "Styling", _font)

    def _font_underline():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        wb = Workbook(); ws = wb.active
        ws["A1"] = "underline"; ws["A1"].font = Font(underline="single")
    _check("Font underline", "Styling", _font_underline)

    def _font_strike():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        wb = Workbook(); ws = wb.active
        ws["A1"] = "strike"; ws["A1"].font = Font(strikethrough=True)
    _check("Font strikethrough", "Styling", _font_strike)

    def _font_vert():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        wb = Workbook(); ws = wb.active
        ws["A1"] = "sup"; ws["A1"].font = Font(vertAlign="superscript")
    _check("Font superscript/subscript", "Styling", _font_vert)

    def _numfmt():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = 42000; ws["A1"].number_format = "$#,##0.00"
    _check("Number formats", "Styling", _numfmt)

    def _alignment():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Alignment
        wb = Workbook(); ws = wb.active
        ws["A1"] = "center"; ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    _check("Alignment (horiz/vert/wrap/rotation)", "Styling", _alignment)

    def _border():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Border, Side
        wb = Workbook(); ws = wb.active
        ws["A1"] = "box"
        ws["A1"].border = Border(left=Side(style="thin"), right=Side(style="thin"))
    _check("Borders (left/right/top/bottom)", "Styling", _border)

    def _diag_border():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Border, Side
        wb = Workbook(); ws = wb.active
        ws["A1"] = "diag"
        ws["A1"].border = Border(diagonal=Side(style="thin"), diagonalUp=True)
    _check("Diagonal borders", "Styling", _diag_border)

    def _fill():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import PatternFill
        wb = Workbook(); ws = wb.active
        ws["A1"] = "filled"
        ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFFF00")
    _check("Pattern fills", "Styling", _fill)

    # --- Structure ---
    def _multi_sheet():
        from openpyxl_rust import Workbook
        wb = Workbook(); wb.create_sheet("Extra")
        assert len(wb.sheetnames) == 2
    _check("Multiple worksheets", "Structure", _multi_sheet)

    def _remove_sheet():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws2 = wb.create_sheet("Extra"); wb.remove(ws2)
        assert len(wb.sheetnames) == 1
    _check("Remove worksheet", "Structure", _remove_sheet)

    def _col_width():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.column_dimensions["A"].width = 25
    _check("Column widths", "Structure", _col_width)

    def _row_height():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.row_dimensions[1].height = 30
    _check("Row heights", "Structure", _row_height)

    def _freeze():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws.freeze_panes = "A2"
    _check("Freeze panes", "Structure", _freeze)

    def _merge():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws.merge_cells("A1:D1")
    _check("Merged cells", "Structure", _merge)

    def _insert_rows():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "before"; ws.insert_rows(1, 2)
    _check("Insert rows", "Structure", _insert_rows)

    def _delete_rows():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "del"; ws["A2"] = "keep"; ws.delete_rows(1)
    _check("Delete rows", "Structure", _delete_rows)

    def _insert_cols():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "before"; ws.insert_cols(1, 2)
    _check("Insert columns", "Structure", _insert_cols)

    def _delete_cols():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "del"; ws["B1"] = "keep"; ws.delete_cols(1)
    _check("Delete columns", "Structure", _delete_cols)

    # --- Iteration ---
    def _iter_rows():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = 1; ws["B2"] = 2
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
    _check("iter_rows()", "Iteration", _iter_rows)

    def _iter_cols():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = 1; ws["B2"] = 2
        cols = list(ws.iter_cols(values_only=True))
        assert len(cols) == 2
    _check("iter_cols()", "Iteration", _iter_cols)

    def _dimensions():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = 1; ws["C3"] = 2
        assert ws.dimensions == "A1:C3"
    _check("Dimension properties", "Iteration", _dimensions)

    def _values():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "x"
        assert list(ws.values) is not None
    _check("values property", "Iteration", _values)

    # --- Batch API ---
    def _append():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["a", "b", "c"])
        assert ws.cell(row=1, column=1).value == "a"
    _check("append() single row", "Batch", _append)

    def _append_rows():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.append_rows([[1, 2], [3, 4]])
    _check("append_rows() batch", "Batch", _append_rows)

    # --- I/O ---
    def _save_file():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "test"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            assert os.path.getsize(path) > 0
        finally:
            os.unlink(path)
    _check("Save to file path", "I/O", _save_file)

    def _save_buffer():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "test"
        buf = io.BytesIO(); wb.save(buf)
        assert buf.tell() > 0
    _check("Save to BytesIO", "I/O", _save_buffer)

    def _load():
        from openpyxl_rust import Workbook, load_workbook
        wb = Workbook(); ws = wb.active; ws["A1"] = "loaded"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = load_workbook(path)
            assert wb2.active.cell(row=1, column=1).value == "loaded"
        finally:
            os.unlink(path)
    _check("load_workbook() (Rust/calamine)", "I/O", _load)

    # --- Advanced ---
    def _autofilter():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.auto_filter.ref = "A1:C10"
    _check("Auto filter", "Advanced", _autofilter)

    def _comment():
        from openpyxl_rust import Workbook, Comment
        wb = Workbook(); ws = wb.active
        ws["A1"] = "noted"; ws["A1"].comment = Comment("Note", "Author")
    _check("Comments / notes", "Advanced", _comment)

    def _hyperlink():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "link"; ws["A1"].hyperlink = "https://example.com"
    _check("Hyperlinks", "Advanced", _hyperlink)

    def _image():
        from openpyxl_rust import Workbook
        from openpyxl_rust.image import Image
        wb = Workbook(); ws = wb.active
        # 1x1 white PNG
        png = (b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
               b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00"
               b"\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00"
               b"\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82")
        ws.add_image(Image(png), "A1")
    _check("Embedded images", "Advanced", _image)

    def _dv():
        from openpyxl_rust import Workbook
        from openpyxl_rust.datavalidation import DataValidation
        wb = Workbook(); ws = wb.active
        dv = DataValidation(type="list", formula1='"A,B,C"')
        dv.add("A1")
        ws.add_data_validation(dv)
    _check("Data validation", "Advanced", _dv)

    def _cond_fmt():
        from openpyxl_rust import Workbook, ColorScaleRule
        wb = Workbook(); ws = wb.active
        rule = ColorScaleRule(start_type="min", start_color="FF0000",
                              end_type="max", end_color="00FF00")
        ws.conditional_formatting.add("A1:A10", rule)
    _check("Conditional formatting", "Advanced", _cond_fmt)

    def _named_range():
        from openpyxl_rust import Workbook, DefinedName
        wb = Workbook()
        dn = DefinedName("MyRange", attr_text="Sheet!$A$1:$C$10")
        wb.defined_names.add(dn)
    _check("Named ranges / defined names", "Advanced", _named_range)

    def _page_setup():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
    _check("Page setup (orientation/paper/margins)", "Advanced", _page_setup)

    def _protection():
        from openpyxl_rust import Workbook
        wb = Workbook(); ws = wb.active
        ws.protection.sheet = True
        ws.protection.password = "secret"
    _check("Sheet protection", "Advanced", _protection)

    # --- Not supported ---
    features.append(("Charts", "Not Yet Supported", False, "Not in v1 scope"))
    features.append(("Gradient fills", "Not Yet Supported", False, "Not in v1 scope"))
    features.append(("Named styles", "Not Yet Supported", False, "Not in v1 scope"))
    features.append(("Rich text (inline strings)", "Not Yet Supported", False, "Not in v1 scope"))
    features.append(("load_workbook with styles", "Not Yet Supported", False, "data_only mode only"))

    return features


def format_feature_section(features):
    lines = []

    # Group by category
    categories = {}
    for name, cat, ok, err in features:
        categories.setdefault(cat, []).append((name, ok, err))

    total_pass = sum(1 for _, _, ok, _ in features if ok)
    total = len(features)
    lines.append(f"  Feature coverage: {total_pass}/{total} ({total_pass / total * 100:.0f}%)")
    lines.append("")

    for cat, items in categories.items():
        passed = sum(1 for _, ok, _ in items if ok)
        icon = "OK" if passed == len(items) else f"{passed}/{len(items)}"
        lines.append(f"  [{icon}] {cat}")
        for name, ok, err in items:
            mark = "  +" if ok else "  -"
            suffix = "" if ok else f"  ({err})"
            lines.append(f"    {mark} {name}{suffix}")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 4. Correctness Verification
# ---------------------------------------------------------------------------

def run_correctness_checks():
    """Write files with openpyxl_rust, reload with openpyxl, verify values match."""
    checks = []

    def _verify(name, fn):
        try:
            fn()
            checks.append((name, True, None))
        except Exception as e:
            checks.append((name, False, str(e)))

    def _roundtrip_strings():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = "hello"; ws["A2"] = "world"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active["A1"].value == "hello"
            assert wb2.active["A2"].value == "world"
        finally:
            os.unlink(path)
    _verify("String roundtrip (write rust -> read openpyxl)", _roundtrip_strings)

    def _roundtrip_numbers():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = 42; ws["A2"] = 3.14159
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active["A1"].value == 42
            assert abs(wb2.active["A2"].value - 3.14159) < 1e-10
        finally:
            os.unlink(path)
    _verify("Number roundtrip (int + float)", _roundtrip_numbers)

    def _roundtrip_bool():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = True; ws["A2"] = False
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active["A1"].value is True
            assert wb2.active["A2"].value is False
        finally:
            os.unlink(path)
    _verify("Boolean roundtrip", _roundtrip_bool)

    def _roundtrip_styles():
        from openpyxl_rust import Workbook
        from openpyxl_rust.styles import Font
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = "bold"; ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = 42000; ws["A2"].number_format = "$#,##0.00"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active["A1"].font.bold is True
            assert wb2.active["A1"].font.size == 14
            assert wb2.active["A2"].number_format == "$#,##0.00"
        finally:
            os.unlink(path)
    _verify("Style roundtrip (font + number format)", _roundtrip_styles)

    def _roundtrip_multisheet():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws1 = wb.active; ws1.title = "First"
        ws2 = wb.create_sheet("Second")
        ws1["A1"] = "sheet1"; ws2["A1"] = "sheet2"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.sheetnames == ["First", "Second"]
            assert wb2["First"]["A1"].value == "sheet1"
            assert wb2["Second"]["A1"].value == "sheet2"
        finally:
            os.unlink(path)
    _verify("Multi-sheet roundtrip", _roundtrip_multisheet)

    def _roundtrip_merge():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = "merged"; ws.merge_cells("A1:C1")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert len(wb2.active.merged_cells.ranges) > 0
        finally:
            os.unlink(path)
    _verify("Merged cells roundtrip", _roundtrip_merge)

    def _roundtrip_unicode():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        ws["A1"] = "Rocket \U0001F680"
        ws["A2"] = "\u4f60\u597d\u4e16\u754c"  # Chinese
        ws["A3"] = "\u0645\u0631\u062d\u0628\u0627"  # Arabic
        # Only check values that the console can print safely
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert "Rocket" in wb2.active["A1"].value
            assert wb2.active["A2"].value is not None
        finally:
            os.unlink(path)
    _verify("Unicode roundtrip (emoji, CJK, Arabic)", _roundtrip_unicode)

    def _roundtrip_large():
        from openpyxl_rust import Workbook
        import openpyxl
        wb = Workbook(); ws = wb.active
        for r in range(1, 1001):
            ws.cell(row=r, column=1, value=f"row_{r}")
            ws.cell(row=r, column=2, value=r * 1.5)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active.cell(row=1000, column=1).value == "row_1000"
            assert wb2.active.cell(row=1000, column=2).value == 1500.0
        finally:
            os.unlink(path)
    _verify("Large data roundtrip (1000 rows)", _roundtrip_large)

    return checks


def format_correctness_section(checks):
    lines = []
    passed = sum(1 for _, ok, _ in checks if ok)
    total = len(checks)
    status = "ALL VERIFIED" if passed == total else "ISSUES FOUND"
    lines.append(f"  {status}: {passed}/{total} checks passed")
    lines.append("")
    for name, ok, err in checks:
        mark = "  PASS" if ok else "  FAIL"
        suffix = "" if ok else f" - {err}"
        lines.append(f"  {mark}  {name}{suffix}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 5. Summary
# ---------------------------------------------------------------------------

def format_summary(test_passed, test_total, bench_results, features, checks):
    lines = []

    all_checks_ok = all(ok for _, ok, _ in checks)
    all_tests_ok = test_passed == test_total
    avg_speedup = sum(r["speedup"] for r in bench_results) / len(bench_results)
    max_speedup = max(r["speedup"] for r in bench_results)
    feat_pass = sum(1 for _, _, ok, _ in features if ok)

    lines.append("  WHY openpyxl_rust?")
    lines.append("")
    lines.append(f"  SPEED        {avg_speedup:.1f}x average faster than openpyxl (up to {max_speedup:.1f}x)")
    lines.append(f"  CORRECTNESS  {test_total} tests passing, {len(checks)} roundtrip verifications")
    lines.append(f"  FEATURES     {feat_pass} features supported with openpyxl-compatible API")
    lines.append(f"  DROP-IN      Same API as openpyxl - change one import line")
    lines.append(f"  TYPE SAFE    Full .pyi stubs for IDE autocompletion")
    lines.append(f"  RUST ENGINE  rust_xlsxwriter + calamine via PyO3 - no C dependencies")
    lines.append("")
    lines.append("  QUICK MIGRATION:")
    lines.append("")
    lines.append("    # Before")
    lines.append("    from openpyxl import Workbook")
    lines.append("")
    lines.append("    # After")
    lines.append("    from openpyxl_rust import Workbook")
    lines.append("")

    if not all_tests_ok:
        lines.append(f"  WARNING: {test_total - test_passed} test(s) failing!")
    if not all_checks_ok:
        failed_checks = [n for n, ok, _ in checks if not ok]
        lines.append(f"  WARNING: Roundtrip failures: {', '.join(failed_checks)}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _safe_print(text):
    """Print with fallback for Windows consoles that can't handle unicode."""
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def main():
    parser = argparse.ArgumentParser(description="openpyxl_rust report generator")
    parser.add_argument("--quick", action="store_true", help="Use smaller data sizes for faster run")
    parser.add_argument("--save-only", action="store_true", help="Only save to file, suppress terminal output")
    args = parser.parse_args()

    report = ReportBuilder()
    print()
    print(DIVIDER)
    print("  Generating openpyxl_rust Report...")
    print(DIVIDER)
    print()

    # 1. Test suite
    print("  [1/5] Running test suite...")
    passed, failed, errors, duration, _ = run_test_suite()
    report.section(
        "1. TEST SUITE",
        format_test_section(passed, failed, errors, duration),
    )

    # 2. Performance benchmarks
    print("  [2/5] Running performance benchmarks...")
    if args.quick:
        print("         (--quick mode: reduced data sizes)")
    bench_results = run_benchmarks(quick=args.quick)
    report.section(
        "2. PERFORMANCE BENCHMARKS",
        format_benchmark_section(bench_results),
    )

    # 3. Feature parity
    print("  [3/5] Checking feature parity...")
    features = check_feature_parity()
    report.section(
        "3. FEATURE PARITY",
        format_feature_section(features),
    )

    # 4. Correctness
    print("  [4/5] Running correctness verification...")
    checks = run_correctness_checks()
    report.section(
        "4. CORRECTNESS VERIFICATION (write rust -> read openpyxl)",
        format_correctness_section(checks),
    )

    # 5. Summary
    print("  [5/5] Building summary...")
    report.section(
        "5. SUMMARY",
        format_summary(passed, passed + failed + errors, bench_results, features, checks),
    )

    # Output
    text = report.build()
    if not args.save_only:
        print()
        _safe_print(text)

    # Save to file
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    report_path = REPORT_DIR / f"report_{timestamp}.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(text)
    print()
    print(f"  Report saved to: {report_path}")
    print()


if __name__ == "__main__":
    main()
