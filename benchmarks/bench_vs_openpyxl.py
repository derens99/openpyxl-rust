# benchmarks/bench_vs_openpyxl.py
"""
Performance comparison: openpyxl_rust vs openpyxl.
Runs identical operations with both libraries and prints a comparison table.
"""

import os
import tempfile
import time


def bench_large_data_openpyxl(path, rows=100_000, cols=10):
    """100k rows x 10 cols of mixed types."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if c % 3 == 0:
                ws.cell(row=r, column=c, value=f"str_{r}_{c}")
            elif c % 3 == 1:
                ws.cell(row=r, column=c, value=r * c * 1.1)
            else:
                ws.cell(row=r, column=c, value=r % 2 == 0)
    wb.save(path)


def bench_large_data_rust(path, rows=100_000, cols=10):
    """100k rows x 10 cols of mixed types."""
    from openpyxl_rust import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if c % 3 == 0:
                ws.cell(row=r, column=c, value=f"str_{r}_{c}")
            elif c % 3 == 1:
                ws.cell(row=r, column=c, value=r * c * 1.1)
            else:
                ws.cell(row=r, column=c, value=r % 2 == 0)
    wb.save(path)


def bench_formatted_openpyxl(path, rows=10_000):
    """10k rows with bold headers, number formats, column widths."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Name", "Revenue", "Cost", "Profit", "Margin"]
    bold = Font(bold=True, size=12)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
    for c in range(1, 6):
        ws.column_dimensions[chr(64 + c)].width = 15
    for r in range(2, rows + 2):
        ws.cell(row=r, column=1, value=f"Item {r}")
        ws.cell(row=r, column=2, value=r * 100.0)
        ws.cell(row=r, column=2).number_format = "$#,##0.00"
        ws.cell(row=r, column=3, value=r * 60.0)
        ws.cell(row=r, column=3).number_format = "$#,##0.00"
        ws.cell(row=r, column=4, value=r * 40.0)
        ws.cell(row=r, column=4).number_format = "$#,##0.00"
        ws.cell(row=r, column=5, value=0.4)
        ws.cell(row=r, column=5).number_format = "0.0%"
    wb.save(path)


def bench_formatted_rust(path, rows=10_000):
    """10k rows with bold headers, number formats, column widths."""
    from openpyxl_rust import Workbook
    from openpyxl_rust.styles import Font

    wb = Workbook()
    ws = wb.active
    headers = ["Name", "Revenue", "Cost", "Profit", "Margin"]
    bold = Font(bold=True, size=12)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
    for c in range(1, 6):
        ws.column_dimensions[chr(64 + c)].width = 15
    for r in range(2, rows + 2):
        ws.cell(row=r, column=1, value=f"Item {r}")
        ws.cell(row=r, column=2, value=r * 100.0)
        ws.cell(row=r, column=2).number_format = "$#,##0.00"
        ws.cell(row=r, column=3, value=r * 60.0)
        ws.cell(row=r, column=3).number_format = "$#,##0.00"
        ws.cell(row=r, column=4, value=r * 40.0)
        ws.cell(row=r, column=4).number_format = "$#,##0.00"
        ws.cell(row=r, column=5, value=0.4)
        ws.cell(row=r, column=5).number_format = "0.0%"
    wb.save(path)


def bench_batch_data_rust(path, rows=100_000, cols=10):
    """100k rows x 10 cols using batch append_rows API."""
    from openpyxl_rust import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    batch = []
    for r in range(1, rows + 1):
        row = []
        for c in range(1, cols + 1):
            if c % 3 == 0:
                row.append(f"str_{r}_{c}")
            elif c % 3 == 1:
                row.append(r * c * 1.1)
            else:
                row.append(r % 2 == 0)
        batch.append(row)
        if len(batch) >= 10_000:
            ws.append_rows(batch)
            batch = []
    if batch:
        ws.append_rows(batch)
    wb.save(path)


def bench_multisheet_openpyxl(path, sheets=5, rows=20_000):
    """5 sheets x 20k rows each."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s in range(sheets):
        ws = wb.create_sheet(f"Sheet{s + 1}")
        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=f"s{s}_r{r}")
            ws.cell(row=r, column=2, value=r * 1.5)
            ws.cell(row=r, column=3, value=r % 2 == 0)
    wb.save(path)


def bench_multisheet_rust(path, sheets=5, rows=20_000):
    """5 sheets x 20k rows each."""
    from openpyxl_rust import Workbook

    wb = Workbook()
    # Rename default sheet as first sheet
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, rows + 1):
        ws.cell(row=r, column=1, value=f"s0_r{r}")
        ws.cell(row=r, column=2, value=r * 1.5)
        ws.cell(row=r, column=3, value=r % 2 == 0)
    for s in range(1, sheets):
        ws = wb.create_sheet(f"Sheet{s + 1}")
        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=f"s{s}_r{r}")
            ws.cell(row=r, column=2, value=r * 1.5)
            ws.cell(row=r, column=3, value=r % 2 == 0)
    wb.save(path)


def run_bench(name, fn_openpyxl, fn_rust):
    """Run a single benchmark pair and return results."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path1 = f.name
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path2 = f.name

    try:
        # openpyxl
        start = time.perf_counter()
        fn_openpyxl(path1)
        t_openpyxl = time.perf_counter() - start
        size_openpyxl = os.path.getsize(path1)

        # openpyxl_rust
        start = time.perf_counter()
        fn_rust(path2)
        t_rust = time.perf_counter() - start
        size_rust = os.path.getsize(path2)

        speedup = t_openpyxl / t_rust if t_rust > 0 else float("inf")

        return {
            "name": name,
            "openpyxl_time": t_openpyxl,
            "rust_time": t_rust,
            "speedup": speedup,
            "openpyxl_size": size_openpyxl,
            "rust_size": size_rust,
        }
    finally:
        os.unlink(path1)
        os.unlink(path2)


def main():
    benchmarks = [
        ("Large data (100k rows x 10 cols)", bench_large_data_openpyxl, bench_large_data_rust),
        ("Batch data (100k rows x 10 cols)", bench_large_data_openpyxl, bench_batch_data_rust),
        ("Formatted (10k rows, styles)", bench_formatted_openpyxl, bench_formatted_rust),
        ("Multi-sheet (5 x 20k rows)", bench_multisheet_openpyxl, bench_multisheet_rust),
    ]

    print("=" * 75)
    print("  openpyxl_rust vs openpyxl - Performance Benchmark")
    print("=" * 75)
    print()

    results = []
    for name, fn_o, fn_r in benchmarks:
        print(f"Running: {name}...")
        r = run_bench(name, fn_o, fn_r)
        results.append(r)

    print()
    print(f"{'Benchmark':<35} {'openpyxl':>10} {'ours':>10} {'speedup':>10}")
    print("-" * 67)
    for r in results:
        print(f"{r['name']:<35} {r['openpyxl_time']:>9.2f}s {r['rust_time']:>9.2f}s {r['speedup']:>9.1f}x")

    print()
    print(f"{'Benchmark':<35} {'openpyxl':>12} {'ours':>12}")
    print("-" * 61)
    for r in results:
        print(f"{r['name']:<35} {r['openpyxl_size'] / 1024:>10.1f} KB {r['rust_size'] / 1024:>10.1f} KB")

    print()
    avg_speedup = sum(r["speedup"] for r in results) / len(results)
    print(f"Average speedup: {avg_speedup:.1f}x faster")
    print()


if __name__ == "__main__":
    main()
