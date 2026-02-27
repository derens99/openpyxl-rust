"""Tests for Worksheet API improvements: range access and unmerge_cells."""

import os
import tempfile

import pytest

from openpyxl_rust import Workbook
from openpyxl_rust.cell import Cell


class TestRangeAccess:
    """Tests for ws['A1:C3'] range access and ws[int] row access."""

    def test_range_access_returns_tuples(self):
        """ws['A1:C3'] returns a tuple of 3 tuples, each with 3 Cell objects."""
        wb = Workbook()
        ws = wb.active
        # Populate a 3x3 grid
        for row in range(1, 4):
            for col in range(1, 4):
                ws.cell(row=row, column=col, value=row * 10 + col)

        result = ws["A1:C3"]
        assert isinstance(result, tuple)
        assert len(result) == 3  # 3 rows
        for row_tuple in result:
            assert isinstance(row_tuple, tuple)
            assert len(row_tuple) == 3  # 3 columns
            for cell in row_tuple:
                assert isinstance(cell, Cell)

    def test_range_access_values(self):
        """Verify that range-accessed cells have the correct values."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["C1"] = "City"
        ws["A2"] = "Alice"
        ws["B2"] = 30
        ws["C2"] = "NYC"
        ws["A3"] = "Bob"
        ws["B3"] = 25
        ws["C3"] = "LA"

        result = ws["A1:C3"]
        # Row 1
        assert result[0][0].value == "Name"
        assert result[0][1].value == "Age"
        assert result[0][2].value == "City"
        # Row 2
        assert result[1][0].value == "Alice"
        assert result[1][1].value == 30
        assert result[1][2].value == "NYC"
        # Row 3
        assert result[2][0].value == "Bob"
        assert result[2][1].value == 25
        assert result[2][2].value == "LA"

    def test_range_access_single_cell_still_works(self):
        """ws['A1'] still returns a single Cell object, not a tuple."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 42

        cell = ws["A1"]
        assert isinstance(cell, Cell)
        assert cell.value == 42

    def test_range_access_row_by_int(self):
        """ws[1] returns a tuple of cells for row 1."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "X"
        ws["B1"] = "Y"
        ws["C1"] = "Z"
        ws["A2"] = 1
        ws["B2"] = 2
        ws["C2"] = 3

        row1 = ws[1]
        assert isinstance(row1, tuple)
        assert len(row1) == 3
        assert row1[0].value == "X"
        assert row1[1].value == "Y"
        assert row1[2].value == "Z"

    def test_range_access_row_by_int_sparse(self):
        """ws[1] on a sparse sheet still returns cells spanning min_column to max_column."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=2, value="B1")
        ws.cell(row=1, column=4, value="D1")

        row1 = ws[1]
        assert isinstance(row1, tuple)
        # Should span from column 2 to column 4 (3 cells)
        assert len(row1) == 3
        assert row1[0].value == "B1"
        assert row1[1].value is None  # C1 created on-demand
        assert row1[2].value == "D1"

    def test_range_access_creates_cells_on_demand(self):
        """Accessing a range that includes empty cells creates them."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "only"

        result = ws["A1:B2"]
        assert len(result) == 2
        assert len(result[0]) == 2
        # A1 has a value, the rest are newly created
        assert result[0][0].value == "only"
        assert result[0][1].value is None
        assert result[1][0].value is None
        assert result[1][1].value is None


class TestUnmergeCells:
    """Tests for ws.unmerge_cells()."""

    def test_unmerge_cells(self):
        """Merge then unmerge; verify ranges list is cleared."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Merged"
        ws.merge_cells("A1:D1")
        assert len(ws.merged_cell_ranges) == 1

        ws.unmerge_cells("A1:D1")
        assert len(ws.merged_cell_ranges) == 0

    def test_unmerge_cells_not_found(self):
        """Unmerging a range that isn't merged raises ValueError."""
        wb = Workbook()
        ws = wb.active

        with pytest.raises(ValueError, match="is not merged"):
            ws.unmerge_cells("A1:D1")

    def test_unmerge_cells_partial(self):
        """Unmerging one range leaves other merges intact."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:D1")
        assert len(ws.merged_cell_ranges) == 2

        ws.unmerge_cells("A1:B1")
        assert len(ws.merged_cell_ranges) == 1
        assert ws.merged_cell_ranges[0] == ("C1", "D1")

    def test_unmerge_cells_save_roundtrip(self):
        """Merge, unmerge, save -- verify openpyxl sees no merge."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        ws.merge_cells("A1:C1")
        ws.unmerge_cells("A1:C1")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            wb.save(tmp_path)

            import openpyxl

            wb2 = openpyxl.load_workbook(tmp_path)
            ws2 = wb2.active
            assert len(ws2.merged_cells.ranges) == 0
        finally:
            os.remove(tmp_path)

    def test_unmerge_cells_case_insensitive(self):
        """Unmerge works regardless of case in the range string."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("a1:d1")  # lowercase merge
        assert len(ws.merged_cell_ranges) == 1

        ws.unmerge_cells("A1:D1")  # uppercase unmerge
        assert len(ws.merged_cell_ranges) == 0
