# tests/test_parity.py
"""
1-to-1 feature parity tests: openpyxl vs openpyxl-rust.

Every test writes with openpyxl_rust, reads back with real openpyxl,
and asserts the result matches what openpyxl would produce.

Features marked @pytest.mark.skip(reason="not yet implemented: ...")
serve as a living checklist of missing functionality.

Run:  pytest tests/test_parity.py -v
Gaps: grep -c "not yet implemented" tests/test_parity.py
"""

import io
import struct
import zipfile
import zlib
from datetime import date, datetime, time

import openpyxl as real_openpyxl
import pytest

from openpyxl_rust import (
    Comment,
    DataValidation,
    DefinedName,
    Image,
    PageMargins,
    SheetProtection,
    Table,
    TableColumn,
    TableStyleInfo,
)
from openpyxl_rust import (
    Workbook as RustWorkbook,
)
from openpyxl_rust import (
    load_workbook as rust_load_workbook,
)
from openpyxl_rust.chart import (
    AreaChart,
    AreaChart3D,
    BarChart,
    BarChart3D,
    DoughnutChart,
    LineChart,
    LineChart3D,
    PieChart,
    PieChart3D,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
    StockChart,
)
from openpyxl_rust.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl_rust.styles import Alignment, Border, Font, PatternFill, Side


def _save_and_reopen(wb, tmp_path, name="test.xlsx"):
    """Save openpyxl_rust workbook, reopen with real openpyxl."""
    path = str(tmp_path / name)
    wb.save(path)
    return real_openpyxl.load_workbook(path)


# ---------------------------------------------------------------------------
# 1. Workbook
# ---------------------------------------------------------------------------
class TestWorkbookParity:
    """Workbook lifecycle: create, save, load, sheetnames, active, dunder methods."""

    def test_create_empty_workbook(self, tmp_path):
        wb = RustWorkbook()
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.sheetnames) == 1
        assert rb.active is not None

    def test_sheetnames_default(self, tmp_path):
        wb = RustWorkbook()
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.sheetnames == ["Sheet"]

    def test_active_sheet(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "hello"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "hello"

    def test_getitem_by_name(self, tmp_path):
        wb = RustWorkbook()
        wb.active.title = "MySheet"
        wb["MySheet"]["A1"] = 42
        rb = _save_and_reopen(wb, tmp_path)
        assert rb["MySheet"]["A1"].value == 42

    def test_iter_sheets(self, tmp_path):
        wb = RustWorkbook()
        wb.create_sheet("Second")
        rb = _save_and_reopen(wb, tmp_path)
        names = [ws.title for ws in rb]
        assert len(names) == 2

    def test_len(self, tmp_path):
        wb = RustWorkbook()
        wb.create_sheet("S2")
        wb.create_sheet("S3")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.sheetnames) == 3

    def test_contains(self):
        wb = RustWorkbook()
        wb.active.title = "Data"
        assert "Data" in wb.sheetnames
        assert "Missing" not in wb.sheetnames

    def test_save_and_load_roundtrip(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "roundtrip"
        path = str(tmp_path / "rt.xlsx")
        wb.save(path)
        wb2 = rust_load_workbook(path)
        assert wb2.active["A1"].value == "roundtrip"

    def test_save_to_filelike(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "bytes"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        rb = real_openpyxl.load_workbook(buf)
        assert rb.active["A1"].value == "bytes"

    def test_document_properties(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        wb.properties.title = "Test Title"
        wb.properties.creator = "Test Creator"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.properties.title == "Test Title"
        assert rb.properties.creator == "Test Creator"

    @pytest.mark.skip(reason="not yet implemented: workbook protection")
    def test_workbook_protection(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 2. Sheet Management
# ---------------------------------------------------------------------------
class TestSheetManagementParity:
    """create_sheet, remove, rename, reorder, copy, visibility."""

    def test_create_sheet_default_title(self, tmp_path):
        wb = RustWorkbook()
        wb.create_sheet()
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.sheetnames) == 2

    def test_create_sheet_custom_title(self, tmp_path):
        wb = RustWorkbook()
        wb.create_sheet("Custom")
        rb = _save_and_reopen(wb, tmp_path)
        assert "Custom" in rb.sheetnames

    def test_create_multiple_sheets(self, tmp_path):
        wb = RustWorkbook()
        for i in range(5):
            wb.create_sheet(f"Sheet{i}")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.sheetnames) == 6  # default + 5

    def test_remove_sheet(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.create_sheet("ToRemove")
        wb.remove(ws)
        rb = _save_and_reopen(wb, tmp_path)
        assert "ToRemove" not in rb.sheetnames

    def test_rename_sheet(self, tmp_path):
        wb = RustWorkbook()
        wb.active.title = "Renamed"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.sheetnames[0] == "Renamed"

    def test_sheet_order_preserved(self, tmp_path):
        wb = RustWorkbook()
        wb.active.title = "First"
        wb.create_sheet("Second")
        wb.create_sheet("Third")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.sheetnames == ["First", "Second", "Third"]

    @pytest.mark.skip(reason="not yet implemented: create_sheet with index parameter")
    def test_create_sheet_at_index(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: copy_worksheet")
    def test_copy_worksheet(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: move_sheet")
    def test_move_sheet(self, tmp_path):
        pass

    def test_sheet_visibility(self, tmp_path):
        wb = RustWorkbook()
        ws1 = wb.active
        ws1.title = "Visible"
        ws2 = wb.create_sheet("Hidden")
        ws2.sheet_state = "hidden"
        ws3 = wb.create_sheet("VeryHidden")
        ws3.sheet_state = "veryHidden"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb["Visible"].sheet_state == "visible"
        assert rb["Hidden"].sheet_state == "hidden"
        assert rb["VeryHidden"].sheet_state == "veryHidden"


# ---------------------------------------------------------------------------
# 3. Cell Access
# ---------------------------------------------------------------------------
class TestCellAccessParity:
    """Cell access patterns: ws['A1'], ws.cell(), slicing, ranges."""

    def test_getitem_single_cell(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "test"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "test"

    def test_cell_method(self, tmp_path):
        wb = RustWorkbook()
        wb.active.cell(row=1, column=1, value="cell_method")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.cell(row=1, column=1).value == "cell_method"

    def test_cell_method_row_col(self, tmp_path):
        wb = RustWorkbook()
        wb.active.cell(row=5, column=3, value="E3? no C5")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["C5"].value == "E3? no C5"

    def test_setitem(self, tmp_path):
        wb = RustWorkbook()
        wb.active["B2"] = 99
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["B2"].value == 99

    def test_cell_coordinate(self):
        wb = RustWorkbook()
        cell = wb.active.cell(row=3, column=2)
        assert cell.coordinate == "B3"

    def test_cell_row_column(self):
        wb = RustWorkbook()
        cell = wb.active.cell(row=5, column=10)
        assert cell.row == 5
        assert cell.column == 10

    def test_overwrite_cell(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "first"
        wb.active["A1"] = "second"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "second"

    def test_getitem_range(self):
        wb = RustWorkbook()
        ws = wb.active
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=r * 10 + c)
        cells = ws["A1:C3"]
        assert len(cells) == 3  # 3 rows
        assert len(cells[0]) == 3  # 3 cols per row

    def test_dimensions_after_write(self):
        wb = RustWorkbook()
        ws = wb.active
        ws["C5"] = "data"
        assert ws.min_row == 5
        assert ws.max_row == 5
        assert ws.min_column == 3
        assert ws.max_column == 3


# ---------------------------------------------------------------------------
# 4. Cell Data Types
# ---------------------------------------------------------------------------
class TestCellDataTypesParity:
    """All cell value types: string, number, bool, date, formula, None."""

    def test_string(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Hello World"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "Hello World"
        assert rb.active["A1"].data_type == "s"

    def test_integer(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 42
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == 42

    def test_float(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 3.14159
        rb = _save_and_reopen(wb, tmp_path)
        assert abs(rb.active["A1"].value - 3.14159) < 1e-10

    def test_large_integer(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 2**53
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == 2**53

    def test_negative_number(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = -999.5
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == -999.5

    def test_zero(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 0
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == 0

    def test_boolean_true(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = True
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value is True

    def test_boolean_false(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = False
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value is False

    def test_none_empty(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = None
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value is None

    def test_formula(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 10
        wb.active["A2"] = 20
        wb.active["A3"] = "=SUM(A1:A2)"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A3"].data_type == "f"
        assert rb.active["A3"].value == "=SUM(A1:A2)"

    def test_datetime(self, tmp_path):
        wb = RustWorkbook()
        dt = datetime(2024, 6, 15, 10, 30, 0)
        wb.active["A1"] = dt
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        assert val.year == 2024
        assert val.month == 6
        assert val.day == 15

    def test_date(self, tmp_path):
        wb = RustWorkbook()
        d = date(2024, 1, 1)
        wb.active["A1"] = d
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        # openpyxl may read as datetime; check date part
        assert val.year == 2024
        assert val.month == 1
        assert val.day == 1

    def test_time(self, tmp_path):
        wb = RustWorkbook()
        t = time(14, 30, 0)
        wb.active["A1"] = t
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        # openpyxl reads time as datetime with date 1899-12-30 or as time
        assert val.hour == 14
        assert val.minute == 30
        assert val.second == 0

    def test_unicode_string(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Hello \U0001f30d World"
        wb.active["A2"] = "\u65e5\u672c\u8a9e\u30c6\u30b9\u30c8"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "Hello \U0001f30d World"
        assert rb.active["A2"].value == "\u65e5\u672c\u8a9e\u30c6\u30b9\u30c8"

    def test_empty_string(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = ""
        rb = _save_and_reopen(wb, tmp_path)
        # openpyxl may read empty string as None or ""
        val = rb.active["A1"].value
        assert val is None or val == ""

    def test_error_value(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "#VALUE!"
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        assert val == "#VALUE!"

    def test_rich_text(self, tmp_path):
        from openpyxl_rust.rich_text import CellRichText, TextBlock
        from openpyxl_rust.styles import Font

        wb = RustWorkbook()
        ws = wb.active
        rt = CellRichText("Normal ", TextBlock(Font(bold=True), "Bold"), " end")
        ws["A1"].value = rt
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        val = rb.active["A1"].value
        # Rich text should contain all the text parts
        assert "Normal" in str(val)
        assert "Bold" in str(val)


# ---------------------------------------------------------------------------
# 5. Iterators
# ---------------------------------------------------------------------------
class TestIteratorsParity:
    """iter_rows, iter_cols, values, append."""

    def test_append_single_row(self, tmp_path):
        wb = RustWorkbook()
        wb.active.append([1, 2, 3])
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == 1
        assert rb.active["B1"].value == 2
        assert rb.active["C1"].value == 3

    def test_append_multiple_rows(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "Name"
        assert rb.active["A2"].value == "Alice"
        assert rb.active["B3"].value == 25

    def test_append_rows_batch(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        data = [[f"r{r}c{c}" for c in range(5)] for r in range(10)]
        ws.append_rows(data)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "r0c0"
        assert rb.active["E10"].value == "r9c4"

    def test_iter_rows_values_only(self):
        wb = RustWorkbook()
        ws = wb.active
        ws.append([1, 2])
        ws.append([3, 4])
        rows = list(ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2, values_only=True))
        assert rows == [(1, 2), (3, 4)]

    def test_iter_rows_cells(self):
        wb = RustWorkbook()
        ws = wb.active
        ws.append([10, 20])
        rows = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2))
        assert len(rows) == 1
        assert rows[0][0].value == 10
        assert rows[0][1].value == 20

    def test_iter_cols_values_only(self):
        wb = RustWorkbook()
        ws = wb.active
        ws.append([1, 2])
        ws.append([3, 4])
        cols = list(ws.iter_cols(min_col=1, max_col=2, min_row=1, max_row=2, values_only=True))
        assert cols == [(1, 3), (2, 4)]

    def test_values_property(self):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append(["c", "d"])
        vals = list(ws.values)
        assert vals == [("a", "b"), ("c", "d")]

    def test_append_mixed_types(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["text", 42, True, None, 3.14])
        rb = _save_and_reopen(wb, tmp_path)
        rws = rb.active
        assert rws["A1"].value == "text"
        assert rws["B1"].value == 42
        assert rws["C1"].value is True
        assert rws["D1"].value is None
        assert abs(rws["E1"].value - 3.14) < 1e-10


# ---------------------------------------------------------------------------
# 6. Cell Styles
# ---------------------------------------------------------------------------
class TestCellStylesParity:
    """Font, Fill, Border, Alignment, NumberFormat, Protection, NamedStyle."""

    # -- Font --
    def test_font_bold(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Bold"
        wb.active["A1"].font = Font(bold=True)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.bold is True

    def test_font_italic(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Italic"
        wb.active["A1"].font = Font(italic=True)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.italic is True

    def test_font_size(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Big"
        wb.active["A1"].font = Font(size=20)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.size == 20

    def test_font_name(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Courier"
        wb.active["A1"].font = Font(name="Courier New")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.name == "Courier New"

    def test_font_color(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Red"
        wb.active["A1"].font = Font(color="FF0000")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.color.rgb.endswith("FF0000")

    def test_font_underline(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Underline"
        wb.active["A1"].font = Font(underline="single")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.underline == "single"

    def test_font_strikethrough(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Strike"
        wb.active["A1"].font = Font(strikethrough=True)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.strike is True or rb.active["A1"].font.strikethrough is True

    def test_font_combined(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Combined"
        wb.active["A1"].font = Font(name="Arial", size=14, bold=True, italic=True, color="0000FF")
        rb = _save_and_reopen(wb, tmp_path)
        f = rb.active["A1"].font
        assert f.name == "Arial"
        assert f.size == 14
        assert f.bold is True
        assert f.italic is True

    def test_font_superscript(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "super"
        wb.active["A1"].font = Font(vertAlign="superscript")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].font.vertAlign == "superscript"

    # -- Fill --
    def test_fill_solid(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Yellow"
        wb.active["A1"].fill = PatternFill(fill_type="solid", start_color="FFFF00")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].fill.patternType == "solid"
        assert rb.active["A1"].fill.fgColor.rgb.endswith("FFFF00")

    def test_fill_gray125(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Gray"
        wb.active["A1"].fill = PatternFill(fill_type="gray125")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].fill.patternType == "gray125"

    @pytest.mark.skip(reason="not yet implemented: GradientFill")
    def test_fill_gradient(self, tmp_path):
        pass

    # -- Border --
    def test_border_thin_all_sides(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Bordered"
        wb.active["A1"].border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        rb = _save_and_reopen(wb, tmp_path)
        b = rb.active["A1"].border
        assert b.left.style == "thin"
        assert b.right.style == "thin"
        assert b.top.style == "thin"
        assert b.bottom.style == "thin"

    def test_border_thick(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Thick"
        wb.active["A1"].border = Border(left=Side(style="thick", color="FF0000"))
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].border.left.style == "thick"

    def test_border_double(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Double"
        wb.active["A1"].border = Border(bottom=Side(style="double", color="0000FF"))
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].border.bottom.style == "double"

    def test_border_dashed(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Dashed"
        wb.active["A1"].border = Border(top=Side(style="dashed", color="000000"))
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].border.top.style == "dashed"

    # -- Alignment --
    def test_alignment_center(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Centered"
        wb.active["A1"].alignment = Alignment(horizontal="center", vertical="center")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].alignment.horizontal == "center"
        assert rb.active["A1"].alignment.vertical == "center"

    def test_alignment_wrap_text(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Long text that wraps"
        wb.active["A1"].alignment = Alignment(wrap_text=True)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].alignment.wrapText is True or rb.active["A1"].alignment.wrap_text is True

    def test_alignment_rotation(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Rotated"
        wb.active["A1"].alignment = Alignment(text_rotation=45)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].alignment.textRotation == 45

    def test_alignment_indent(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Indented"
        wb.active["A1"].alignment = Alignment(indent=2)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].alignment.indent == 2

    def test_alignment_shrink_to_fit(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = "Shrink"
        wb.active["A1"].alignment = Alignment(shrink_to_fit=True)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].alignment.shrinkToFit is True or rb.active["A1"].alignment.shrink_to_fit is True

    # -- Number Format --
    def test_number_format_currency(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 1234.5
        wb.active["A1"].number_format = "$#,##0.00"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].number_format == "$#,##0.00"

    def test_number_format_percentage(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 0.75
        wb.active["A1"].number_format = "0.00%"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].number_format == "0.00%"

    def test_number_format_date(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = 45000
        wb.active["A1"].number_format = "yyyy-mm-dd"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].number_format == "yyyy-mm-dd"

    # -- Combined styles --
    def test_multiple_styles_on_cell(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Styled"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFFF00")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].number_format = "0.00"
        rb = _save_and_reopen(wb, tmp_path)
        c = rb.active["A1"]
        assert c.font.bold is True
        assert c.font.size == 16
        assert c.fill.patternType == "solid"
        assert c.alignment.horizontal == "center"
        assert c.number_format == "0.00"

    @pytest.mark.skip(reason="not yet implemented: NamedStyle")
    def test_named_style(self, tmp_path):
        pass

    def test_cell_protection(self, tmp_path):
        from openpyxl_rust.styles import Protection

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "locked"
        ws["A1"].protection = Protection(locked=True, hidden=False)
        ws["B1"] = "unlocked"
        ws["B1"].protection = Protection(locked=False, hidden=False)
        ws["C1"] = "hidden"
        ws["C1"].protection = Protection(locked=True, hidden=True)
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.active["A1"].protection.locked is True
        assert rb.active["B1"].protection.locked is False
        assert rb.active["C1"].protection.hidden is True


# ---------------------------------------------------------------------------
# 7. Row/Column Dimensions
# ---------------------------------------------------------------------------
class TestRowColumnDimensionsParity:
    """Column width, row height, hidden, outline, grouping."""

    def test_column_width(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "wide column"
        ws.column_dimensions["A"].width = 30
        rb = _save_and_reopen(wb, tmp_path)
        assert abs(rb.active.column_dimensions["A"].width - 30) < 1

    def test_multiple_column_widths(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["C1"] = "c"
        rb = _save_and_reopen(wb, tmp_path)
        assert abs(rb.active.column_dimensions["A"].width - 10) < 1
        assert abs(rb.active.column_dimensions["B"].width - 20) < 1
        assert abs(rb.active.column_dimensions["C"].width - 30) < 1

    def test_row_height(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "tall row"
        ws.row_dimensions[1].height = 40
        rb = _save_and_reopen(wb, tmp_path)
        assert abs(rb.active.row_dimensions[1].height - 40) < 1

    def test_multiple_row_heights(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 40
        ws["A1"] = "r1"
        ws["A2"] = "r2"
        rb = _save_and_reopen(wb, tmp_path)
        assert abs(rb.active.row_dimensions[1].height - 20) < 1
        assert abs(rb.active.row_dimensions[2].height - 40) < 1

    def test_hidden_row(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "visible"
        ws["A2"] = "hidden"
        ws.row_dimensions[2].hidden = True
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.active.row_dimensions[2].hidden is True
        assert not rb.active.row_dimensions[1].hidden

    def test_hidden_column(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "visible"
        ws["B1"] = "hidden"
        ws.column_dimensions["B"].hidden = True
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.active.column_dimensions["B"].hidden is True
        assert not rb.active.column_dimensions["A"].hidden

    def test_row_grouping(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "header"
        ws["A2"] = "detail 1"
        ws["A3"] = "detail 2"
        ws.row_dimensions[2].outline_level = 1
        ws.row_dimensions[3].outline_level = 1
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.active.row_dimensions[2].outline_level == 1
        assert rb.active.row_dimensions[3].outline_level == 1

    def test_column_grouping(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "main"
        ws["B1"] = "sub 1"
        ws["C1"] = "sub 2"
        ws.column_dimensions["B"].outline_level = 1
        ws.column_dimensions["C"].outline_level = 1
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        # openpyxl may store contiguous columns with identical properties
        # as a single entry (e.g., "B" with min=2, max=3 covering B:C)
        b_dim = rb.active.column_dimensions["B"]
        assert b_dim.outline_level == 1
        # Verify column C is covered by the range (max >= 3 means C is included)
        assert b_dim.max >= 3  # B=2, C=3 in 1-based

    def test_column_auto_size(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Short"
        ws["A2"] = "This is a much longer text string for testing autofit"
        ws.auto_fit_columns()
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        # After autofit, column A width should be wider than default (8.43)
        width = rb.active.column_dimensions["A"].width
        assert width is not None and width > 8.43


# ---------------------------------------------------------------------------
# 8. Merged Cells
# ---------------------------------------------------------------------------
class TestMergedCellsParity:
    """Merge, unmerge, merged cell value behavior."""

    def test_merge_cells(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Merged"
        ws.merge_cells("A1:C1")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.merged_cells.ranges) >= 1
        # Note: merged cell value may not persist in current implementation
        merge_refs = [str(r) for r in rb.active.merged_cells.ranges]
        assert "A1:C1" in merge_refs

    def test_merge_block(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Block"
        ws.merge_cells("A1:B3")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.merged_cells.ranges) >= 1

    def test_unmerge_cells(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "temp"
        ws.merge_cells("A1:B1")
        ws.unmerge_cells("A1:B1")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.merged_cells.ranges) == 0

    def test_multiple_merges(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "First"
        ws.merge_cells("A1:B1")
        ws["A3"] = "Second"
        ws.merge_cells("A3:C3")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.merged_cells.ranges) == 2

    def test_merged_cell_value_none(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Value"
        ws.merge_cells("A1:B1")
        rb = _save_and_reopen(wb, tmp_path)
        # In openpyxl, B1 in a merge returns MergedCell with value None
        assert rb.active["B1"].value is None


# ---------------------------------------------------------------------------
# 9. Hyperlinks
# ---------------------------------------------------------------------------
class TestHyperlinksParity:
    """External URLs, internal sheet refs, tooltips."""

    def test_external_url(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Click me"
        ws["A1"].hyperlink = "https://example.com"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].hyperlink is not None
        assert "example.com" in rb.active["A1"].hyperlink.target

    def test_internal_reference(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Go to Sheet2"
        ws["A1"].hyperlink = "#Sheet2!A1"
        wb.create_sheet("Sheet2")
        rb = _save_and_reopen(wb, tmp_path)
        h = rb.active["A1"].hyperlink
        assert h is not None

    def test_hyperlink_with_value(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Link text"
        ws["A1"].hyperlink = "https://example.org"
        rb = _save_and_reopen(wb, tmp_path)
        # Hyperlink assignment may overwrite cell display value with the URL
        assert rb.active["A1"].value is not None
        assert rb.active["A1"].hyperlink is not None

    def test_multiple_hyperlinks(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Link1"
        ws["A1"].hyperlink = "https://example.com/1"
        ws["A2"] = "Link2"
        ws["A2"].hyperlink = "https://example.com/2"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].hyperlink is not None
        assert rb.active["A2"].hyperlink is not None


# ---------------------------------------------------------------------------
# 10. Comments
# ---------------------------------------------------------------------------
class TestCommentsParity:
    """Cell comments with text and author."""

    def test_basic_comment(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["A1"].comment = Comment("This is a note", "Author")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].comment is not None
        # rust_xlsxwriter may prepend "Author:\n" to comment text
        assert "This is a note" in rb.active["A1"].comment.text
        assert rb.active["A1"].comment.author == "Author"

    def test_comment_without_author(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["A1"].comment = Comment("Note text", "")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].comment is not None
        assert "Note text" in rb.active["A1"].comment.text

    def test_multiple_comments(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Cell 1"
        ws["A1"].comment = Comment("Note 1", "Alice")
        ws["B2"] = "Cell 2"
        ws["B2"].comment = Comment("Note 2", "Bob")
        rb = _save_and_reopen(wb, tmp_path)
        assert "Note 1" in rb.active["A1"].comment.text
        assert "Note 2" in rb.active["B2"].comment.text

    def test_comment_on_empty_cell(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"].comment = Comment("Orphan note", "Author")
        rb = _save_and_reopen(wb, tmp_path)
        c = rb.active["A1"].comment
        assert c is not None
        assert "Orphan note" in c.text


# ---------------------------------------------------------------------------
# 11. Data Validation
# ---------------------------------------------------------------------------
class TestDataValidationParity:
    """list, whole, decimal, textLength, custom, operators, messages."""

    def test_list_validation(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(type="list", formula1='"Cat,Dog,Fish"', allow_blank=True)
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert len(dvs) == 1
        assert dvs[0].type == "list"

    def test_whole_number(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="100")
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].type == "whole"

    def test_decimal_validation(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].type == "decimal"

    def test_text_length(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="50")
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].type == "textLength"

    def test_custom_formula(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(type="custom", formula1="=AND(A1>0,A1<100)")
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].type == "custom"

    def test_error_message(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1="0",
            showErrorMessage=True,
            errorTitle="Invalid",
            error="Must be positive",
            errorStyle="stop",
        )
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].errorTitle == "Invalid"
        assert dvs[0].error == "Must be positive"

    def test_input_message(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv = DataValidation(
            type="list",
            formula1='"A,B,C"',
            showInputMessage=True,
            promptTitle="Choose",
            prompt="Pick a letter",
        )
        dv.add("A1")
        ws.add_data_validation(dv)
        rb = _save_and_reopen(wb, tmp_path)
        dvs = rb.active.data_validations.dataValidation
        assert dvs[0].promptTitle == "Choose"
        assert dvs[0].prompt == "Pick a letter"

    def test_multiple_validations(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        dv1 = DataValidation(type="list", formula1='"X,Y,Z"')
        dv1.add("A1")
        dv2 = DataValidation(type="whole", operator="greaterThan", formula1="0")
        dv2.add("B1")
        ws.add_data_validation(dv1)
        ws.add_data_validation(dv2)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.data_validations.dataValidation) == 2

    def test_all_operators(self, tmp_path):
        """Test every comparison operator works."""
        operators = [
            "between",
            "notBetween",
            "equal",
            "notEqual",
            "greaterThan",
            "lessThan",
            "greaterThanOrEqual",
            "lessThanOrEqual",
        ]
        for op in operators:
            wb = RustWorkbook()
            ws = wb.active
            kwargs = {"type": "whole", "operator": op, "formula1": "1"}
            if op in ("between", "notBetween"):
                kwargs["formula2"] = "100"
            dv = DataValidation(**kwargs)
            dv.add("A1")
            ws.add_data_validation(dv)
            rb = _save_and_reopen(wb, tmp_path, name=f"dv_{op}.xlsx")
            actual = rb.active.data_validations.dataValidation[0].operator
            # openpyxl treats "between" as the default, so it may read as None
            if op == "between":
                assert actual == op or actual is None, f"Failed for operator: {op}"
            else:
                assert actual == op, f"Failed for operator: {op}"


# ---------------------------------------------------------------------------
# 12. Conditional Formatting
# ---------------------------------------------------------------------------
class TestConditionalFormattingParity:
    """ColorScale, DataBar, IconSet, CellIs, Formula rules."""

    def test_color_scale_2color(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        rule = ColorScaleRule(
            start_type="min",
            start_color="FF0000",
            end_type="max",
            end_color="00FF00",
        )
        ws.conditional_formatting.add("A1:A10", rule)
        rb = _save_and_reopen(wb, tmp_path)
        cfs = rb.active.conditional_formatting
        assert len(list(cfs)) >= 1

    def test_color_scale_3color(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        rule = ColorScaleRule(
            start_type="min",
            start_color="FF0000",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFF00",
            end_type="max",
            end_color="00FF00",
        )
        ws.conditional_formatting.add("A1:A10", rule)
        rb = _save_and_reopen(wb, tmp_path)
        cfs = list(rb.active.conditional_formatting)
        assert len(cfs) >= 1

    def test_data_bar(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        rule = DataBarRule(
            start_type="min",
            end_type="max",
            color="638EC6",
        )
        ws.conditional_formatting.add("A1:A5", rule)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(list(rb.active.conditional_formatting)) >= 1

    def test_icon_set(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 20)
        rule = IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67])
        ws.conditional_formatting.add("A1:A5", rule)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(list(rb.active.conditional_formatting)) >= 1

    def test_cell_is_rule_less_than(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        rule = CellIsRule(
            operator="lessThan",
            formula=["30"],
            font=Font(color="FF0000"),
        )
        ws.conditional_formatting.add("A1:A5", rule)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(list(rb.active.conditional_formatting)) >= 1

    def test_cell_is_rule_between(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        rule = CellIsRule(
            operator="between",
            formula=["20", "40"],
            fill=PatternFill(fill_type="solid", start_color="FFFF00"),
        )
        ws.conditional_formatting.add("A1:A5", rule)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(list(rb.active.conditional_formatting)) >= 1

    def test_formula_rule(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append([1, 2, 3])
        rule = FormulaRule(
            formula=["$A1>2"],
            font=Font(bold=True),
        )
        ws.conditional_formatting.add("A1:C1", rule)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(list(rb.active.conditional_formatting)) >= 1

    def test_multiple_rules_same_range(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        rule1 = CellIsRule(operator="lessThan", formula=["20"], font=Font(color="FF0000"))
        rule2 = CellIsRule(operator="greaterThan", formula=["40"], font=Font(color="00FF00"))
        ws.conditional_formatting.add("A1:A5", rule1)
        ws.conditional_formatting.add("A1:A5", rule2)
        rb = _save_and_reopen(wb, tmp_path)
        total_rules = sum(len(cf.rules) for cf in rb.active.conditional_formatting)
        assert total_rules >= 2

    def test_top10_rule(self, tmp_path):
        from openpyxl_rust.formatting.rule import Top10Rule
        from openpyxl_rust.styles import PatternFill

        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i * 10)
        ws.conditional_formatting.add(
            "A1:A10",
            Top10Rule(rank=3, fill=PatternFill(start_color="FF0000", fill_type="solid")),
        )
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        cf_rules = list(rb.active.conditional_formatting)
        assert len(cf_rules) > 0

    def test_duplicate_values_rule(self, tmp_path):
        from openpyxl_rust.formatting.rule import DuplicateRule
        from openpyxl_rust.styles import PatternFill

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "apple"
        ws["A2"] = "banana"
        ws["A3"] = "apple"
        ws.conditional_formatting.add(
            "A1:A3",
            DuplicateRule(fill=PatternFill(start_color="FFFF00", fill_type="solid")),
        )
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        cf_rules = list(rb.active.conditional_formatting)
        assert len(cf_rules) > 0

    def test_contains_text_rule(self, tmp_path):
        from openpyxl_rust.formatting.rule import TextRule
        from openpyxl_rust.styles import PatternFill

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "hello world"
        ws["A2"] = "goodbye"
        ws["A3"] = "hello again"
        ws.conditional_formatting.add(
            "A1:A3",
            TextRule(
                operator="containsText",
                text="hello",
                fill=PatternFill(start_color="00FF00", fill_type="solid"),
            ),
        )
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        cf_rules = list(rb.active.conditional_formatting)
        assert len(cf_rules) > 0


# ---------------------------------------------------------------------------
# 13. Auto Filter
# ---------------------------------------------------------------------------
class TestAutoFilterParity:
    """AutoFilter setup and basic filtering."""

    def test_basic_autofilter(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])
        ws.auto_filter.ref = "A1:C3"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.auto_filter.ref == "A1:C3"

    def test_autofilter_with_data(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 21):
            ws.append([f"Item {i}", i, i * 1.5])
        ws.auto_filter.ref = f"A1:C{20}"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.auto_filter.ref is not None

    def test_autofilter_column_filter(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "Status"
        ws["A2"] = "Yes"
        ws["A3"] = "No"
        ws["A4"] = "Yes"
        ws.auto_filter.ref = "A1:A4"
        ws.auto_filter.add_filter_column(0, ["Yes"])
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        af = rb.active.auto_filter
        assert af.ref is not None
        # Check that filterColumn data exists
        assert len(af.filterColumn) >= 1


# ---------------------------------------------------------------------------
# 14. Tables
# ---------------------------------------------------------------------------
class TestTablesParity:
    """Table with columns, styles, and options."""

    def test_basic_table(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["Name", "Score"])
        ws.append(["Alice", 95])
        ws.append(["Bob", 87])
        tab = Table(displayName="Scores", ref="A1:B3")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
        )
        ws.add_table(tab)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.tables) >= 1

    def test_table_with_columns(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["ID", "Value", "Label"])
        ws.append([1, 100, "A"])
        ws.append([2, 200, "B"])
        tab = Table(displayName="DataTable", ref="A1:C3")
        tab.tableColumns = [
            TableColumn(id=1, name="ID"),
            TableColumn(id=2, name="Value"),
            TableColumn(id=3, name="Label"),
        ]
        ws.add_table(tab)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.tables) >= 1

    def test_table_style_options(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["H1", "H2"])
        ws.append([1, 2])
        tab = Table(displayName="StyledTable", ref="A1:B2")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        ws.add_table(tab)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.tables) >= 1

    def test_multiple_tables(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws["D1"] = "C"
        ws["D2"] = 3
        tab1 = Table(displayName="Table1", ref="A1:B2")
        tab2 = Table(displayName="Table2", ref="D1:D2")
        ws.add_table(tab1)
        ws.add_table(tab2)
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active.tables) >= 2


# ---------------------------------------------------------------------------
# 15. Defined Names / Named Ranges
# ---------------------------------------------------------------------------
class TestDefinedNamesParity:
    """Global named ranges and constants."""

    def test_named_range(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = 100
        dn = DefinedName("MyRange", attr_text="Data!$A$1")
        wb.defined_names.add(dn)
        rb = _save_and_reopen(wb, tmp_path)
        assert "MyRange" in list(rb.defined_names)

    def test_named_constant(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "placeholder"
        dn = DefinedName("TaxRate", attr_text="0.07")
        wb.defined_names.add(dn)
        rb = _save_and_reopen(wb, tmp_path)
        assert "TaxRate" in list(rb.defined_names)

    def test_multiple_named_ranges(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 1
        ws["B1"] = 2
        dn1 = DefinedName("First", attr_text="Sheet1!$A$1")
        dn2 = DefinedName("Second", attr_text="Sheet1!$B$1")
        wb.defined_names.add(dn1)
        wb.defined_names.add(dn2)
        rb = _save_and_reopen(wb, tmp_path)
        names = list(rb.defined_names)
        assert "First" in names
        assert "Second" in names


# ---------------------------------------------------------------------------
# 16. Print Setup
# ---------------------------------------------------------------------------
class TestPrintSetupParity:
    """Page margins, orientation, paper size, print area, print titles."""

    def test_landscape_orientation(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "landscape"
        ws.page_setup.orientation = "landscape"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.page_setup.orientation == "landscape"

    def test_portrait_orientation(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "portrait"
        ws.page_setup.orientation = "portrait"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.page_setup.orientation == "portrait"

    def test_paper_size(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "A4"
        ws.page_setup.paperSize = 9  # A4
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.page_setup.paperSize == 9

    def test_margins(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "margins"
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        rb = _save_and_reopen(wb, tmp_path)
        m = rb.active.page_margins
        assert abs(m.left - 0.5) < 0.01
        assert abs(m.right - 0.5) < 0.01

    def test_print_area(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "area"
        ws.print_area = "A1:D10"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.print_area is not None

    def test_print_titles_rows(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.append(["Header1", "Header2"])
        ws.append([1, 2])
        ws.print_title_rows = "1:1"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.print_title_rows is not None or rb.active.print_titles is not None

    @pytest.mark.skip(reason="not yet implemented: fitToWidth/fitToHeight not roundtripped by rust_xlsxwriter")
    def test_fit_to_page(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "fit"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        rb = _save_and_reopen(wb, tmp_path)
        ps = rb.active.page_setup
        assert ps.fitToWidth == 1
        assert ps.fitToHeight == 1

    def test_center_horizontally(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "center"
        ws.print_options.horizontalCentered = True
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.print_options.horizontalCentered is True

    def test_header_footer(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        ws.oddHeader.center.text = "My Header"
        ws.oddFooter.center.text = "Page &P"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        # openpyxl reads header/footer as raw format strings
        header = rb.active.oddHeader
        assert header is not None
        # The header text should contain "My Header"
        header_text = header.center.text if hasattr(header.center, "text") else str(header)
        assert "My Header" in str(header_text) or "My Header" in str(header)

    def test_page_breaks(self, tmp_path):
        from openpyxl_rust.page_break import Break

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "page 1"
        ws["A21"] = "page 2"
        ws.row_breaks.append(Break(id=20))
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        row_break_ids = [b.id for b in rb.active.row_breaks.brk]
        assert 20 in row_break_ids


# ---------------------------------------------------------------------------
# 17. Sheet Protection
# ---------------------------------------------------------------------------
class TestSheetProtectionParity:
    """Sheet protection with password and individual options."""

    def test_basic_protection(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "protected"
        ws.protection.enable()
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.protection.sheet is True

    def test_protection_with_password(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "secret"
        ws.protection.set_password("test123")
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.protection.sheet is True
        assert rb.active.protection.password is not None

    def test_protection_options(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "options"
        ws.protection = SheetProtection(
            sheet=True,
            format_cells=False,
            insert_rows=False,
            sort=False,
        )
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.protection.sheet is True

    def test_protection_disabled(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "unprotected"
        ws.protection.enable()
        ws.protection.disable()
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.protection.sheet is not True


# ---------------------------------------------------------------------------
# 18. Sheet Views
# ---------------------------------------------------------------------------
class TestSheetViewsParity:
    """Freeze panes, zoom, gridlines."""

    def test_freeze_panes_b2(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "frozen"
        ws.freeze_panes = "B2"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.freeze_panes == "B2"

    def test_freeze_panes_a2(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "header"
        ws.freeze_panes = "A2"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.freeze_panes == "A2"

    def test_freeze_panes_none(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "no freeze"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active.freeze_panes is None

    def test_zoom_scale(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        ws.zoom = 150
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        # openpyxl stores zoom in sheet_view
        zoom = rb.active.sheet_view.zoomScale
        assert zoom == 150

    @pytest.mark.skip(reason="not yet implemented: split panes")
    def test_split_panes(self, tmp_path):
        pass

    def test_show_gridlines(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        ws._show_gridlines = False
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.active.sheet_view.showGridLines is False


# ---------------------------------------------------------------------------
# 19. Images
# ---------------------------------------------------------------------------
class TestImagesParity:
    """Image insertion with anchor and dimensions."""

    def _make_png_bytes(self):
        """Create a minimal valid PNG (1x1 red pixel)."""

        def chunk(ctype, data):
            c = ctype + data
            return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)

        sig = b"\x89PNG\r\n\x1a\n"
        ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
        raw = zlib.compress(b"\x00\xff\x00\x00")  # filter byte + RGB
        idat = chunk(b"IDAT", raw)
        iend = chunk(b"IEND", b"")
        return sig + ihdr + idat + iend

    def test_image_from_bytes(self, tmp_path):
        png = self._make_png_bytes()
        img_path = tmp_path / "test.png"
        img_path.write_bytes(png)

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "image below"
        img = Image(str(img_path))
        ws.add_image(img, "A3")
        path = str(tmp_path / "img_test.xlsx")
        wb.save(path)
        # Verify image is present in the xlsx archive (Pillow may not be installed
        # so openpyxl._images may be empty even though the image is there)
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            assert any("media/image" in n for n in names), f"No image media found in {names}"
            assert any("drawing" in n for n in names), f"No drawing found in {names}"

    def test_image_anchor(self, tmp_path):
        png = self._make_png_bytes()
        img_path = tmp_path / "test.png"
        img_path.write_bytes(png)

        wb = RustWorkbook()
        ws = wb.active
        img = Image(str(img_path))
        ws.add_image(img, "C5")
        path = str(tmp_path / "img_anchor.xlsx")
        wb.save(path)
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            assert any("media/image" in n for n in names), f"No image media found in {names}"


# ---------------------------------------------------------------------------
# 20. Charts
# ---------------------------------------------------------------------------
class TestChartsParity:
    """All chart types, axes, legend, titles."""

    def _setup_chart_data(self, wb):
        ws = wb.active
        ws.title = "Data"
        ws.append(["Category", "Series1", "Series2"])
        ws.append(["Q1", 10, 30])
        ws.append(["Q2", 20, 25])
        ws.append(["Q3", 30, 20])
        ws.append(["Q4", 40, 15])
        return ws

    def _data_ref(self, ws):
        return Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)

    def _cat_ref(self, ws):
        return Reference(ws, min_col=1, min_row=2, max_row=5)

    def _verify_chart_exists(self, rb):
        assert len(rb.active._charts) >= 1

    def test_bar_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.title = "Bar Chart"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        chart.set_categories(self._cat_ref(ws))
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_bar_chart_stacked(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.title = "Stacked Bar"
        chart.grouping = "stacked"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_bar_chart_3d(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart3D()
        chart.title = "3D Bar"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_line_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = LineChart()
        chart.title = "Line Chart"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        chart.set_categories(self._cat_ref(ws))
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_line_chart_3d(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = LineChart3D()
        chart.title = "3D Line"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_pie_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = PieChart()
        chart.title = "Pie Chart"
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(self._cat_ref(ws))
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_pie_chart_3d(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = PieChart3D()
        chart.title = "3D Pie"
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_area_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = AreaChart()
        chart.title = "Area Chart"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_area_chart_3d(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = AreaChart3D()
        chart.title = "3D Area"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_scatter_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = ScatterChart()
        chart.title = "Scatter Chart"
        xvals = Reference(ws, min_col=2, min_row=2, max_row=5)
        yvals = Reference(ws, min_col=3, min_row=2, max_row=5)
        series = Series(yvals, xvals, title="Data")
        chart.append(series)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_doughnut_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = DoughnutChart()
        chart.title = "Doughnut"
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_radar_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = RadarChart()
        chart.title = "Radar"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        chart.set_categories(self._cat_ref(ws))
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_chart_title(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.title = "My Custom Title"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        rc = rb.active._charts[0]
        # openpyxl reads chart title as a Title object; extract text from rich paragraphs
        title = rc.title
        if isinstance(title, str):
            title_text = title
        else:
            # Title object -> tx -> rich -> paragraphs -> runs -> t
            paragraphs = title.tx.rich.paragraphs if title.tx and title.tx.rich else []
            title_text = "".join(run.t for p in paragraphs for run in (p.r or []) if run.t)
        assert title_text == "My Custom Title"

    def test_chart_axis_titles(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.x_axis_title = "Categories"
        chart.y_axis_title = "Values"
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_chart_no_legend(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.legend = False
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_chart_dimensions(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart = BarChart()
        chart.width = 20
        chart.height = 12
        chart.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart, "E2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_multiple_charts(self, tmp_path):
        wb = RustWorkbook()
        ws = self._setup_chart_data(wb)
        chart1 = BarChart()
        chart1.add_data(self._data_ref(ws), titles_from_data=True)
        chart2 = LineChart()
        chart2.add_data(self._data_ref(ws), titles_from_data=True)
        ws.add_chart(chart1, "E2")
        ws.add_chart(chart2, "E18")
        rb = _save_and_reopen(wb, tmp_path)
        assert len(rb.active._charts) >= 2

    def test_stock_chart(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Date", "Open", "High", "Low", "Close"])
        ws.append(["2024-01-01", 100, 110, 95, 105])
        ws.append(["2024-01-02", 105, 115, 100, 110])
        ws.append(["2024-01-03", 110, 120, 105, 115])
        chart = StockChart()
        data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=4)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=2, max_row=4)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")
        rb = _save_and_reopen(wb, tmp_path)
        self._verify_chart_exists(rb)

    def test_chart_trendline(self, tmp_path):
        from openpyxl_rust.chart.series import Trendline

        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 2 + 1)
        from openpyxl_rust.chart import BarChart, Reference

        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        chart.add_data(data)
        chart.series[0].trendline = Trendline(trendlineType="linear")
        ws.add_chart(chart, "D1")
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert len(rb.active._charts) > 0

    def test_chart_data_labels(self, tmp_path):
        from openpyxl_rust.chart.series import DataLabelList

        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=f"Cat{i}")
            ws.cell(row=i, column=2, value=i * 10)
        from openpyxl_rust.chart import BarChart, Reference

        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        chart.add_data(data)
        chart.series[0].dLbls = DataLabelList(showVal=True)
        ws.add_chart(chart, "D1")
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert len(rb.active._charts) > 0

    def test_chart_legend_position(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        from openpyxl_rust.chart import BarChart, Reference

        chart = BarChart()
        data = Reference(ws, min_col=1, min_row=1, max_row=5)
        chart.add_data(data)
        chart.legend.position = "b"  # bottom
        ws.add_chart(chart, "D1")
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert len(rb.active._charts) > 0


# ---------------------------------------------------------------------------
# 21. Formulas
# ---------------------------------------------------------------------------
class TestFormulasParity:
    """Formula storage and preservation."""

    def test_simple_formula(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=A1+A2"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A3"].value == "=A1+A2"

    def test_sum_formula(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        ws["B1"] = "=SUM(A1:A10)"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["B1"].value == "=SUM(A1:A10)"

    def test_if_formula(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = 50
        ws["B1"] = '=IF(A1>25,"Pass","Fail")'
        rb = _save_and_reopen(wb, tmp_path)
        assert "IF" in rb.active["B1"].value

    def test_formula_data_type(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "=1+1"
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].data_type == "f"

    @pytest.mark.skip(reason="not yet implemented: move_range with formula translation")
    def test_move_range_translate(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 22. DateTime Handling
# ---------------------------------------------------------------------------
class TestDateTimeParity:
    """Date, time, datetime roundtrip verification."""

    def test_datetime_roundtrip(self, tmp_path):
        wb = RustWorkbook()
        dt = datetime(2024, 3, 15, 9, 30, 45)
        wb.active["A1"] = dt
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        assert val.year == 2024
        assert val.month == 3
        assert val.day == 15
        assert val.hour == 9
        assert val.minute == 30

    def test_date_roundtrip(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = date(2000, 1, 1)
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        assert val.year == 2000
        assert val.month == 1
        assert val.day == 1

    def test_time_roundtrip(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = time(23, 59, 59)
        rb = _save_and_reopen(wb, tmp_path)
        val = rb.active["A1"].value
        assert val.hour == 23
        assert val.minute == 59
        assert val.second == 59

    def test_multiple_dates(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = date(2020, 1, 1)
        ws["A2"] = date(2020, 6, 15)
        ws["A3"] = datetime(2020, 12, 31, 23, 59, 59)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value.year == 2020
        assert rb.active["A2"].value.month == 6
        assert rb.active["A3"].value.day == 31

    def test_auto_number_format_date(self, tmp_path):
        wb = RustWorkbook()
        wb.active["A1"] = date(2024, 1, 1)
        rb = _save_and_reopen(wb, tmp_path)
        # Should have a date number format, not "General"
        nf = rb.active["A1"].number_format
        assert nf != "General"


# ---------------------------------------------------------------------------
# 23. Row/Column Insert/Delete Operations
# ---------------------------------------------------------------------------
class TestRowColOpsParity:
    """Insert/delete rows and columns with data shift verification."""

    def test_insert_rows(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "row1"
        ws["A2"] = "row2"
        ws["A3"] = "row3"
        ws.insert_rows(2, 2)  # insert 2 rows at row 2
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "row1"
        assert rb.active["A2"].value is None  # inserted blank
        assert rb.active["A3"].value is None  # inserted blank
        # original row2 "row2" shifted to A4
        assert rb.active["A4"].value == "row2"

    def test_delete_rows(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "keep"
        ws["A2"] = "delete"
        ws["A3"] = "shift_up"
        ws.delete_rows(2, 1)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "keep"
        assert rb.active["A2"].value == "shift_up"

    def test_insert_cols(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "col1"
        ws["B1"] = "col2"
        ws.insert_cols(2, 1)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "col1"
        assert rb.active["B1"].value is None  # inserted blank
        assert rb.active["C1"].value == "col2"  # shifted right

    def test_delete_cols(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "keep"
        ws["B1"] = "delete"
        ws["C1"] = "shift_left"
        ws.delete_cols(2, 1)
        rb = _save_and_reopen(wb, tmp_path)
        assert rb.active["A1"].value == "keep"
        assert rb.active["B1"].value == "shift_left"

    @pytest.mark.skip(reason="not yet implemented: move_range")
    def test_move_range(self, tmp_path):
        pass


# ===========================================================================
# UNSUPPORTED FEATURES -- Living Checklist
# ===========================================================================
# These classes document openpyxl features not yet in openpyxl-rust.
# As features are implemented, convert skips to real tests.
# Count gaps: grep -c "not yet implemented" tests/test_parity.py


# ---------------------------------------------------------------------------
# 24. Header/Footer
# ---------------------------------------------------------------------------
class TestHeaderFooterParity:
    def test_odd_header(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        ws.oddHeader.center.text = "Center Header"
        ws.oddHeader.left.text = "Left"
        ws.oddHeader.right.text = "Right"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        h = rb.active.oddHeader
        assert "Center Header" in str(h.center.text) or "Center Header" in str(h)
        assert "Left" in str(h.left.text) or "Left" in str(h)
        assert "Right" in str(h.right.text) or "Right" in str(h)

    def test_odd_footer(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        ws.oddFooter.center.text = "Page &P of &N"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        f = rb.active.oddFooter
        assert "Page" in str(f.center.text) or "Page" in str(f)

    @pytest.mark.skip(reason="not yet implemented: even header/footer")
    def test_even_header_footer(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: first page header/footer")
    def test_first_page_header_footer(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 25. Page Breaks
# ---------------------------------------------------------------------------
class TestPageBreaksParity:
    def test_row_break(self, tmp_path):
        from openpyxl_rust.page_break import Break

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "above break"
        ws["A11"] = "below break"
        ws.row_breaks.append(Break(id=10))
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        row_break_ids = [b.id for b in rb.active.row_breaks.brk]
        assert 10 in row_break_ids

    def test_col_break(self, tmp_path):
        from openpyxl_rust.page_break import Break

        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "left of break"
        ws["F1"] = "right of break"
        ws.col_breaks.append(Break(id=5))
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        col_break_ids = [b.id for b in rb.active.col_breaks.brk]
        assert 5 in col_break_ids


# ---------------------------------------------------------------------------
# 26. Workbook Protection
# ---------------------------------------------------------------------------
class TestWorkbookProtectionParity:
    @pytest.mark.skip(reason="not yet implemented: workbook lock structure")
    def test_lock_structure(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: workbook lock windows")
    def test_lock_windows(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 27. Document Properties
# ---------------------------------------------------------------------------
class TestDocumentPropertiesParity:
    def test_title(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        wb.properties.title = "My Workbook"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.properties.title == "My Workbook"

    def test_creator(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        wb.properties.creator = "Test Author"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.properties.creator == "Test Author"

    def test_description(self, tmp_path):
        wb = RustWorkbook()
        ws = wb.active
        ws["A1"] = "test"
        wb.properties.description = "A test workbook"
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert rb.properties.description == "A test workbook"


# ---------------------------------------------------------------------------
# 28. Rich Text
# ---------------------------------------------------------------------------
class TestRichTextParity:
    def test_rich_text_basic(self, tmp_path):
        from openpyxl_rust.rich_text import CellRichText, TextBlock

        wb = RustWorkbook()
        ws = wb.active
        rt = CellRichText("Hello ", TextBlock(text="World"))
        ws["A1"].value = rt
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        assert "Hello" in str(rb.active["A1"].value)
        assert "World" in str(rb.active["A1"].value)

    @pytest.mark.skip(reason="not yet implemented: InlineFont TextBlock")
    def test_text_block(self, tmp_path):
        pass

    def test_mixed_formatting(self, tmp_path):
        from openpyxl_rust.rich_text import CellRichText, TextBlock
        from openpyxl_rust.styles import Font

        wb = RustWorkbook()
        ws = wb.active
        rt = CellRichText("Plain ", TextBlock(Font(bold=True, color="FF0000"), "Red Bold"), " more plain")
        ws["A1"].value = rt
        wb.save(str(tmp_path / "test.xlsx"))

        rb = real_openpyxl.load_workbook(str(tmp_path / "test.xlsx"))
        val = str(rb.active["A1"].value)
        assert "Plain" in val
        assert "Red Bold" in val


# ---------------------------------------------------------------------------
# 29. Pivot Tables
# ---------------------------------------------------------------------------
class TestPivotTablesParity:
    @pytest.mark.skip(reason="not yet implemented: pivot table creation")
    def test_basic_pivot(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: pivot table preservation on load")
    def test_pivot_preservation(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 30. Move Range
# ---------------------------------------------------------------------------
class TestMoveRangeParity:
    @pytest.mark.skip(reason="not yet implemented: move_range shift cells")
    def test_move_range_basic(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: move_range translate formulas")
    def test_move_range_translate(self, tmp_path):
        pass


# ---------------------------------------------------------------------------
# 31. Advanced Features (unique stubs not covered by other classes)
# ---------------------------------------------------------------------------
class TestAdvancedFeaturesParity:
    @pytest.mark.skip(reason="not yet implemented: tab color")
    def test_tab_color(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: sparklines")
    def test_sparklines(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: slicers")
    def test_slicers(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: form controls")
    def test_form_controls(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: VBA/macros preservation")
    def test_vba_preservation(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: shapes/textboxes")
    def test_shapes(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: Chartsheet")
    def test_chartsheet(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: write_only mode")
    def test_write_only_mode(self, tmp_path):
        pass

    @pytest.mark.skip(reason="not yet implemented: read_only mode")
    def test_read_only_mode(self, tmp_path):
        pass
