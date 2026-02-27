# tests/test_comments.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Comment, Workbook


def test_comment_basic():
    """Set cell.comment = Comment(...), save, verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["A1"].comment = Comment("This is a comment", "Author")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].comment is not None
        # Note: rust_xlsxwriter embeds author in the VML comment text as "Author:\nText"
        assert "This is a comment" in sheet["A1"].comment.text
    finally:
        os.unlink(path)


def test_comment_with_author():
    """Verify author is set correctly."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    ws["A1"].comment = Comment("Note text", "John Doe")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].comment is not None
        assert "Note text" in sheet["A1"].comment.text
        assert sheet["A1"].comment.author == "John Doe"
    finally:
        os.unlink(path)


def test_comment_multiple_cells():
    """Comments on several cells."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Cell 1"
    ws["A1"].comment = Comment("Comment 1", "Author A")
    ws["B2"] = "Cell 2"
    ws["B2"].comment = Comment("Comment 2", "Author B")
    ws["C3"] = "Cell 3"
    ws["C3"].comment = Comment("Comment 3", "Author C")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].comment is not None
        assert "Comment 1" in sheet["A1"].comment.text
        assert sheet["A1"].comment.author == "Author A"
        assert sheet["B2"].comment is not None
        assert "Comment 2" in sheet["B2"].comment.text
        assert sheet["B2"].comment.author == "Author B"
        assert sheet["C3"].comment is not None
        assert "Comment 3" in sheet["C3"].comment.text
        assert sheet["C3"].comment.author == "Author C"
    finally:
        os.unlink(path)


def test_comment_none_author():
    """Comment with None author - no author set."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "No author"
    ws["A1"].comment = Comment("text without author", None)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].comment is not None
        assert "text without author" in sheet["A1"].comment.text
    finally:
        os.unlink(path)


def test_comment_save_valid():
    """Save with comments, load with openpyxl, verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Data"
    ws["A1"].comment = Comment("Important note", "Reviewer")
    ws["A2"] = 42
    ws["A2"].comment = Comment("Numeric cell note", "Reviewer")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].comment is not None
        assert "Important note" in sheet["A1"].comment.text
        assert sheet["A1"].comment.author == "Reviewer"
        assert sheet["A2"].comment is not None
        assert "Numeric cell note" in sheet["A2"].comment.text
        assert sheet["A2"].comment.author == "Reviewer"
    finally:
        os.unlink(path)
