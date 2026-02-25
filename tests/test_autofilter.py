# tests/test_autofilter.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Workbook


def test_autofilter_basic():
    """Set ws.auto_filter.ref = 'A1:D10', save, verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["C1"] = "Header3"
    ws["D1"] = "Header4"
    ws.auto_filter.ref = "A1:D10"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.auto_filter.ref is not None
        assert sheet.auto_filter.ref == "A1:D10"
    finally:
        os.unlink(path)


def test_autofilter_with_data():
    """Populate data + set filter, verify file is valid."""
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ["Name", "Age", "City", "Score"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    data = [
        ["Alice", 30, "NYC", 95],
        ["Bob", 25, "LA", 88],
        ["Charlie", 35, "Chicago", 72],
        ["Diana", 28, "Boston", 91],
    ]
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.auto_filter.ref = "A1:D5"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # Verify data is intact
        assert sheet["A1"].value == "Name"
        assert sheet["B2"].value == 30
        # Verify autofilter
        assert sheet.auto_filter.ref == "A1:D5"
    finally:
        os.unlink(path)


def test_autofilter_save_load():
    """Save + load with openpyxl, check auto_filter.ref is present."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Col1"
    ws["B1"] = "Col2"
    ws["A2"] = 1
    ws["B2"] = 2
    ws["A3"] = 3
    ws["B3"] = 4
    ws.auto_filter.ref = "A1:B3"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        # Load with openpyxl and verify
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.auto_filter.ref is not None
        assert "A1" in sheet.auto_filter.ref
        assert "B3" in sheet.auto_filter.ref
    finally:
        os.unlink(path)


def test_autofilter_none():
    """No filter set, verify save works fine."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["A2"] = "World"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].value == "Hello"
        assert sheet["A2"].value == "World"
        # No autofilter should be set
        assert sheet.auto_filter.ref is None
    finally:
        os.unlink(path)
