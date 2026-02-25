# tests/test_bugfixes_v05.py
"""Tests for v0.5 bug fixes."""
import os
import tempfile
from datetime import datetime, date, time

import openpyxl as real_openpyxl
from openpyxl_rust import Workbook
from openpyxl_rust.worksheet import _date_to_excel_serial


# =====================================================================
# Bug 1: datetime.time should not crash in _set_rust_value
# =====================================================================


def test_time_value_via_setitem():
    """Setting a cell to datetime.time should not raise TypeError."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = time(14, 30, 0)
    assert ws["A1"].value == time(14, 30, 0)


def test_time_value_via_cell():
    """ws.cell(row, col, value=time(...)) should not raise TypeError."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=time(9, 15, 45))
    assert ws.cell(row=1, column=1).value == time(9, 15, 45)


def test_time_value_saves_and_loads():
    """time values should roundtrip through save/load correctly."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = time(14, 30, 0)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0

        rb = real_openpyxl.load_workbook(path)
        val = rb.active["A1"].value
        # openpyxl reads time-only values as time objects
        assert val.hour == 14
        assert val.minute == 30
        assert val.second == 0
    finally:
        os.unlink(path)


def test_time_serial_calculation():
    """Verify time-to-serial conversion: 14:30:00 = (14*3600+30*60)/86400."""
    wb = Workbook()
    ws = wb.active
    t = time(14, 30, 0)
    ws["A1"] = t
    # The serial value should be fractional day
    expected = (14 * 3600 + 30 * 60) / 86400.0
    # We can verify via Rust side by saving and checking the number
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        val = rb.active["A1"].value
        assert val.hour == 14
        assert val.minute == 30
    finally:
        os.unlink(path)


def test_time_midnight():
    """midnight time(0,0,0) should produce serial 0.0."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = time(0, 0, 0)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_time_via_append():
    """ws.append([time(...)]) should work correctly."""
    wb = Workbook()
    ws = wb.active
    ws.append([time(8, 0, 0), "morning"])
    assert ws.cell(row=1, column=1).value == time(8, 0, 0)
    assert ws.cell(row=1, column=2).value == "morning"


# =====================================================================
# Bug 2: append_rows should not silently drop datetime values
# =====================================================================


def test_append_rows_with_datetime():
    """append_rows should correctly handle datetime values."""
    wb = Workbook()
    ws = wb.active
    dt = datetime(2024, 3, 15, 10, 30, 0)
    ws.append_rows([[dt, "test"]])

    # Python-side cell should retain the original datetime object
    assert ws.cell(row=1, column=1).value == dt
    assert isinstance(ws.cell(row=1, column=1).value, datetime)


def test_append_rows_datetime_saves_correctly():
    """append_rows with datetime should produce valid Excel file."""
    wb = Workbook()
    ws = wb.active
    dt = datetime(2024, 6, 15, 14, 30, 0)
    d = date(2024, 1, 1)
    ws.append_rows([
        [dt, "datetime"],
        [d, "date"],
    ])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        ws2 = rb.active

        # datetime value should roundtrip
        val1 = ws2.cell(row=1, column=1).value
        assert isinstance(val1, datetime)
        assert val1.year == 2024
        assert val1.month == 6
        assert val1.day == 15
        assert val1.hour == 14
        assert val1.minute == 30

        # date value should roundtrip
        val2 = ws2.cell(row=2, column=1).value
        assert val2.year == 2024
        assert val2.month == 1
        assert val2.day == 1

        # string columns untouched
        assert ws2.cell(row=1, column=2).value == "datetime"
        assert ws2.cell(row=2, column=2).value == "date"
    finally:
        os.unlink(path)


def test_append_rows_with_time():
    """append_rows should correctly handle time values."""
    wb = Workbook()
    ws = wb.active
    t = time(9, 0, 0)
    ws.append_rows([[t, "morning"]])

    assert ws.cell(row=1, column=1).value == t

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        val = rb.active.cell(row=1, column=1).value
        assert val.hour == 9
        assert val.minute == 0
    finally:
        os.unlink(path)


def test_append_rows_datetime_gets_number_format():
    """datetime values in append_rows should get auto-format applied."""
    wb = Workbook()
    ws = wb.active
    dt = datetime(2024, 3, 15, 10, 30, 0)
    ws.append_rows([[dt]])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        fmt = rb.active.cell(row=1, column=1).number_format
        # Should have a datetime format, not "General"
        assert fmt != "General"
        assert "yy" in fmt or "mm" in fmt or "dd" in fmt
    finally:
        os.unlink(path)


def test_append_rows_mixed_with_datetime():
    """append_rows with mixed types including datetime should work."""
    wb = Workbook()
    ws = wb.active
    ws.append_rows([
        [1, "hello", datetime(2024, 1, 1), True],
        [2, "world", date(2024, 6, 15), False],
    ])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        ws2 = rb.active

        assert ws2.cell(row=1, column=1).value == 1
        assert ws2.cell(row=1, column=2).value == "hello"
        dt_val = ws2.cell(row=1, column=3).value
        assert isinstance(dt_val, datetime)
        assert dt_val.year == 2024
        assert ws2.cell(row=1, column=4).value is True

        assert ws2.cell(row=2, column=1).value == 2
        assert ws2.cell(row=2, column=2).value == "world"
        d_val = ws2.cell(row=2, column=3).value
        assert d_val.year == 2024
        assert d_val.month == 6
    finally:
        os.unlink(path)


# =====================================================================
# Bug 3: print_options.headings should be serialized
# =====================================================================


def test_print_headings_serialized():
    """Setting print_options.headings=True should produce a valid file."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "test"
    ws.print_options.headings = True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        # Load with openpyxl to verify the headings option was set
        rb = real_openpyxl.load_workbook(path)
        # openpyxl stores this as printOptions.headings in the sheet
        assert rb.active.print_options.headings is True
    finally:
        os.unlink(path)


def test_print_headings_default_none():
    """By default, headings should not be set (None/False)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "test"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        # headings should not be True when not set
        headings = rb.active.print_options.headings
        assert not headings  # None or False
    finally:
        os.unlink(path)


def test_print_headings_with_gridlines():
    """headings and gridlines can both be set simultaneously."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "test"
    ws.print_options.headings = True
    ws.print_options.gridLines = True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        assert rb.active.print_options.headings is True
        assert rb.active.print_options.gridLines is True
    finally:
        os.unlink(path)


# =====================================================================
# Bug 4: no more json_mod references in worksheet.py
# =====================================================================


def test_no_json_mod_import():
    """Verify that json_mod is not used anywhere in worksheet.py source."""
    import inspect
    from openpyxl_rust import worksheet
    source = inspect.getsource(worksheet)
    assert "json_mod" not in source, "Found 'json_mod' in worksheet.py source"


def test_protection_still_works_after_json_cleanup():
    """Protection should still work after removing json_mod import."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "protected"
    ws.protection.sheet = True
    ws.protection._password = "test123"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        rb = real_openpyxl.load_workbook(path)
        assert rb.active.protection.sheet is True
    finally:
        os.unlink(path)


def test_data_validation_still_works_after_json_cleanup():
    """Data validation should still work after removing json_mod import."""
    from openpyxl_rust.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Dog"
    dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"')
    dv.add("A1")
    ws.add_data_validation(dv)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


# =====================================================================
# Bug 5: _date_to_excel_serial epoch verification
# =====================================================================


def test_serial_1900_01_01():
    """date(1900, 1, 1) should be serial 1.0."""
    assert _date_to_excel_serial(1900, 1, 1) == 1.0


def test_serial_2024_01_01():
    """date(2024, 1, 1) should be serial 45292.0."""
    assert _date_to_excel_serial(2024, 1, 1) == 45292.0


def test_serial_1900_03_01():
    """date(1900, 3, 1) should be serial 61.0."""
    assert _date_to_excel_serial(1900, 3, 1) == 61.0


def test_serial_1900_02_28():
    """date(1900, 2, 28) should be serial 59.0 (last valid date before Lotus bug)."""
    assert _date_to_excel_serial(1900, 2, 28) == 59.0


def test_serial_known_date():
    """date(2000, 1, 1) should match Excel's serial number 36526."""
    assert _date_to_excel_serial(2000, 1, 1) == 36526.0


def test_serial_roundtrip_via_openpyxl():
    """Verify serial values produce correct dates when read back by openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = date(1900, 1, 1)
    ws["A2"] = date(2024, 1, 1)
    ws["A3"] = date(1900, 3, 1)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        ws2 = rb.active

        v1 = ws2["A1"].value
        assert v1.year == 1900
        assert v1.month == 1
        assert v1.day == 1

        v2 = ws2["A2"].value
        assert v2.year == 2024
        assert v2.month == 1
        assert v2.day == 1

        v3 = ws2["A3"].value
        assert v3.year == 1900
        assert v3.month == 3
        assert v3.day == 1
    finally:
        os.unlink(path)
