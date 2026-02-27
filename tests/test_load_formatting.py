"""Tests for formatting-preserving load (data_only=False)."""
import os
import tempfile
import pytest
import openpyxl
from openpyxl.styles import Font as OFont, PatternFill as OFill, Border as OBorder, Side as OSide, Alignment as OAlign
from openpyxl_rust import load_workbook, Workbook
from openpyxl_rust.styles import Font, PatternFill, Border, Side, Alignment


def _create_formatted_xlsx(path):
    """Create a test xlsx file with various formatting using real openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formatted"

    # Bold + red font
    ws["A1"] = "Bold Red"
    ws["A1"].font = OFont(name="Arial", size=14, bold=True, color="FF0000")

    # Italic + underline
    ws["A2"] = "Italic Underline"
    ws["A2"].font = OFont(italic=True, underline="single")

    # Yellow fill
    ws["B1"] = "Yellow Fill"
    ws["B1"].fill = OFill(fill_type="solid", fgColor="FFFF00")

    # Borders
    ws["B2"] = "Bordered"
    ws["B2"].border = OBorder(
        left=OSide(style="thin", color="000000"),
        right=OSide(style="thick", color="FF0000"),
        top=OSide(style="double", color="0000FF"),
        bottom=OSide(style="thin", color="000000"),
    )

    # Alignment
    ws["C1"] = "Centered"
    ws["C1"].alignment = OAlign(horizontal="center", vertical="center", wrap_text=True)

    # Number format
    ws["C2"] = 1234.5678
    ws["C2"].number_format = "#,##0.00"

    # Column width
    ws.column_dimensions["A"].width = 25

    # Row height
    ws.row_dimensions[1].height = 30

    # Merge cells
    ws.merge_cells("D1:E2")
    ws["D1"] = "Merged"

    # Freeze panes
    ws.freeze_panes = "B2"

    # Second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Sheet2 Data"
    ws2["A1"].font = OFont(bold=True, size=16)

    wb.save(path)


class TestLoadWithFormatting:
    def test_data_only_true_still_works(self, tmp_path):
        """data_only=True should still use the fast calamine path."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Bold Red"
        assert ws.cell(row=1, column=2).value == "Yellow Fill"

    def test_load_preserves_values(self, tmp_path):
        """data_only=False loads all cell values."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        assert ws.title == "Formatted"
        assert ws.cell(row=1, column=1).value == "Bold Red"
        assert ws.cell(row=1, column=2).value == "Yellow Fill"
        assert ws.cell(row=2, column=1).value == "Italic Underline"
        assert ws.cell(row=2, column=2).value == "Bordered"
        assert ws.cell(row=1, column=3).value == "Centered"
        assert abs(ws.cell(row=2, column=3).value - 1234.5678) < 0.001
        assert ws.cell(row=1, column=4).value == "Merged"

    def test_load_preserves_bold_font(self, tmp_path):
        """data_only=False preserves bold font."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font is not None
        assert cell.font.bold is True
        assert cell.font.name == "Arial"
        assert cell.font.size == 14

    def test_load_preserves_font_color(self, tmp_path):
        """data_only=False preserves font color."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font is not None
        assert cell.font.color is not None
        assert cell.font.color.upper() == "FF0000"

    def test_load_preserves_italic_underline(self, tmp_path):
        """data_only=False preserves italic and underline."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        assert cell.font is not None
        assert cell.font.italic is True
        assert cell.font.underline == "single"

    def test_load_preserves_fill(self, tmp_path):
        """data_only=False preserves fill."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=1, column=2)
        assert cell.fill is not None
        assert cell.fill.fill_type == "solid"
        assert cell.fill.start_color is not None
        assert cell.fill.start_color.upper() == "FFFF00"

    def test_load_preserves_borders(self, tmp_path):
        """data_only=False preserves border styles."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        assert cell.border is not None
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thick"
        assert cell.border.top.style == "double"

    def test_load_preserves_alignment(self, tmp_path):
        """data_only=False preserves alignment."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=1, column=3)
        assert cell.alignment is not None
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.alignment.wrap_text is True

    def test_load_preserves_number_format(self, tmp_path):
        """data_only=False preserves number format."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        cell = ws.cell(row=2, column=3)
        assert cell.number_format == "#,##0.00"

    def test_load_preserves_freeze_panes(self, tmp_path):
        """data_only=False preserves freeze panes."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        assert ws.freeze_panes == "B2"

    def test_load_preserves_merged_cells(self, tmp_path):
        """data_only=False preserves merged cells."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        assert len(ws.merged_cell_ranges) > 0
        # Verify the merge range
        merge_strs = [str(r) for r in ws.merged_cell_ranges]
        assert any("D1" in m and "E2" in m for m in merge_strs)

    def test_load_preserves_multiple_sheets(self, tmp_path):
        """data_only=False loads all sheets."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "Formatted"
        assert wb.sheetnames[1] == "Sheet2"

        ws2 = wb["Sheet2"]
        assert ws2.cell(row=1, column=1).value == "Sheet2 Data"
        assert ws2.cell(row=1, column=1).font.bold is True
        assert ws2.cell(row=1, column=1).font.size == 16

    def test_round_trip_with_formatting(self, tmp_path):
        """Load with formatting, save, and re-load — verify formatting survives."""
        src_path = str(tmp_path / "source.xlsx")
        _create_formatted_xlsx(src_path)

        # Load with formatting
        wb = load_workbook(src_path, data_only=False)

        # Save
        dst_path = str(tmp_path / "dest.xlsx")
        wb.save(dst_path)

        # Re-load with openpyxl to verify
        wb2 = openpyxl.load_workbook(dst_path)
        ws2 = wb2.active

        # Check bold font
        assert ws2["A1"].font.bold is True

        # Check fill
        assert ws2["B1"].fill.fgColor is not None

        # Check border
        assert ws2["B2"].border.left.style == "thin"

        # Check alignment
        assert ws2["C1"].alignment.horizontal == "center"

        # Check number format
        assert ws2["C2"].number_format == "#,##0.00"

        # Check values
        assert ws2["A1"].value == "Bold Red"
        assert abs(ws2["C2"].value - 1234.5678) < 0.001

    def test_load_column_width(self, tmp_path):
        """data_only=False preserves column widths."""
        path = str(tmp_path / "test.xlsx")
        _create_formatted_xlsx(path)

        wb = load_workbook(path, data_only=False)
        ws = wb.active
        assert ws.column_dimensions["A"].width == 25
