"""Tests for Chart support."""
import os
import tempfile
import pytest
import openpyxl
from openpyxl_rust import Workbook
from openpyxl_rust.chart import (
    Reference, Series,
    BarChart, LineChart, PieChart, AreaChart,
    ScatterChart, DoughnutChart, RadarChart,
)


def _make_data_ws(wb):
    """Helper: create a worksheet with sample data for charting."""
    ws = wb.active
    ws.title = "Sheet1"
    # Headers
    ws.append(["Category", "Series1", "Series2"])
    # Data rows
    ws.append(["Q1", 10, 30])
    ws.append(["Q2", 20, 25])
    ws.append(["Q3", 30, 20])
    ws.append(["Q4", 40, 15])
    return ws


class TestBarChart:
    def test_basic_bar_chart(self, tmp_path):
        """Create a bar chart with data and verify xlsx is valid."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.title = "Quarterly Sales"
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "bar_chart.xlsx")
        wb.save(path)

        # Verify with openpyxl
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        assert len(ws2._charts) == 1

    def test_bar_chart_horizontal(self, tmp_path):
        """BarChart with type='bar' produces horizontal bars."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.type = "bar"  # horizontal
        chart.title = "Horizontal"
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        chart.add_data(data)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "bar_horiz.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestLineChart:
    def test_line_chart(self, tmp_path):
        """Create a line chart."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = LineChart()
        chart.title = "Trend"
        chart.y_axis_title = "Value"
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "line_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestPieChart:
    def test_pie_chart(self, tmp_path):
        """Create a pie chart."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = PieChart()
        chart.title = "Distribution"
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "pie_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestAreaChart:
    def test_area_chart(self, tmp_path):
        """Create an area chart."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = AreaChart()
        chart.title = "Area"
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "area_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestScatterChart:
    def test_scatter_chart(self, tmp_path):
        """Create a scatter chart with explicit X/Y series."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["X", "Y"])
        ws.append([1.0, 2.5])
        ws.append([2.0, 4.0])
        ws.append([3.0, 3.5])
        ws.append([4.0, 5.0])

        chart = ScatterChart()
        chart.title = "XY Plot"
        chart.x_axis_title = "X values"
        chart.y_axis_title = "Y values"

        x_values = Reference(ws, min_col=1, min_row=2, max_row=5)
        y_values = Reference(ws, min_col=2, min_row=2, max_row=5)
        s = Series(values=y_values, categories=x_values, title="Data")
        chart.append(s)

        ws.add_chart(chart, "D2")

        path = str(tmp_path / "scatter_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestDoughnutChart:
    def test_doughnut_chart(self, tmp_path):
        """Create a doughnut chart."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = DoughnutChart()
        chart.title = "Donut"
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        chart.add_data(data)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "doughnut_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestRadarChart:
    def test_radar_chart(self, tmp_path):
        """Create a radar chart."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = RadarChart()
        chart.title = "Radar"
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "radar_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1


class TestChartOptions:
    def test_chart_no_legend(self, tmp_path):
        """Chart with legend disabled."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.title = "No Legend"
        chart.legend = False
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        chart.add_data(data)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "no_legend.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1

    def test_chart_axis_titles(self, tmp_path):
        """Chart with axis titles."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.title = "With Axes"
        chart.x_axis_title = "Quarter"
        chart.y_axis_title = "Revenue ($)"
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "axis_titles.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1

    def test_chart_custom_size(self, tmp_path):
        """Chart with custom dimensions."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.title = "Big Chart"
        chart.width = 20   # cm
        chart.height = 12  # cm
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        chart.add_data(data)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "big_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1

    def test_multiple_charts_on_sheet(self, tmp_path):
        """Multiple charts on one sheet."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart1 = BarChart()
        chart1.title = "Bar"
        data1 = Reference(ws, min_col=2, min_row=2, max_row=5)
        chart1.add_data(data1)
        ws.add_chart(chart1, "E2")

        chart2 = LineChart()
        chart2.title = "Line"
        data2 = Reference(ws, min_col=3, min_row=2, max_row=5)
        chart2.add_data(data2)
        ws.add_chart(chart2, "E18")

        path = str(tmp_path / "multi_chart.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 2

    def test_chart_titles_from_data(self, tmp_path):
        """add_data with titles_from_data=True extracts series names from first row."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")

        # Should have 2 series (one per data column: cols 2 and 3)
        assert len(chart.series) == 2

        path = str(tmp_path / "titles_from_data.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1

    def test_stacked_bar_chart(self, tmp_path):
        """BarChart with stacked grouping."""
        wb = Workbook()
        ws = _make_data_ws(wb)

        chart = BarChart()
        chart.grouping = "stacked"
        chart.title = "Stacked"
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")

        path = str(tmp_path / "stacked_bar.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert len(wb2.active._charts) == 1
