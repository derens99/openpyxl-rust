# tests/test_save_styles.py
import os
import tempfile
from openpyxl_rust import Workbook
from openpyxl_rust.styles import Font, Alignment, Border, Side, PatternFill


def _save_and_check(setup_fn):
    wb = Workbook()
    setup_fn(wb)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        return path
    except Exception:
        os.unlink(path)
        raise


def test_save_alignment():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Centered"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    path = _save_and_check(setup)
    os.unlink(path)


def test_save_border():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Bordered"
        ws["A1"].border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
    path = _save_and_check(setup)
    os.unlink(path)


def test_save_fill():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Yellow"
        ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFFF00")
    path = _save_and_check(setup)
    os.unlink(path)


def test_save_all_styles_combined():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Styled"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].border = Border(left=Side(style="thick", color="FF0000"))
        ws["A1"].fill = PatternFill(fill_type="solid", start_color="00FF00")
        ws["A1"].number_format = "#,##0.00"
    path = _save_and_check(setup)
    os.unlink(path)


def test_compat_alignment():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Centered"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        cell = rb.active["A1"]
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"
        assert cell.alignment.wrap_text is True
    finally:
        os.unlink(path)


def test_compat_border():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Bordered"
    ws["A1"].border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        cell = rb.active["A1"]
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
    finally:
        os.unlink(path)


def test_compat_fill():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yellow"
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFFF00")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        cell = rb.active["A1"]
        assert cell.fill.patternType == "solid"
        assert "FFFF00" in cell.fill.start_color.rgb
    finally:
        os.unlink(path)


def test_compat_font_strikethrough():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Strike"
    ws["A1"].font = Font(strikethrough=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        assert rb.active["A1"].font.strikethrough is True
    finally:
        os.unlink(path)


def test_compat_font_superscript():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Super"
    ws["A1"].font = Font(vertAlign="superscript")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        assert rb.active["A1"].font.vertAlign == "superscript"
    finally:
        os.unlink(path)


def test_compat_border_diagonal():
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Diag"
    ws["A1"].border = Border(
        diagonal=Side(style="thin"),
        diagonalUp=True
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        cell = rb.active["A1"]
        assert cell.border.diagonal.border_style == "thin"
    finally:
        os.unlink(path)


def test_batch_format_many_cells():
    """Verify batch format flush works with many formatted cells."""
    import os, tempfile
    from openpyxl_rust import Workbook
    from openpyxl_rust.styles import Font
    wb = Workbook()
    ws = wb.active
    bold = Font(bold=True)
    for r in range(1, 101):
        ws.cell(row=r, column=1, value=f"row{r}")
        ws.cell(row=r, column=1).font = bold
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        import openpyxl
        rb = openpyxl.load_workbook(path)
        assert rb.active.cell(row=1, column=1).font.bold is True
        assert rb.active.cell(row=100, column=1).font.bold is True
    finally:
        os.unlink(path)
