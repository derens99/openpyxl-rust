# tests/test_batch.py
import os
import tempfile

import openpyxl

from openpyxl_rust import Workbook


def test_append_single_row():
    """ws.append([1, 'hello', True]) stores cells at row 1."""
    wb = Workbook()
    ws = wb.active
    ws.append([1, "hello", True])

    # _cells uses 1-based (row, col) keys
    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=1, column=2).value == "hello"
    assert ws.cell(row=1, column=3).value is True


def test_append_multiple_rows():
    """Multiple ws.append() calls produce consecutive rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["c", "d"])
    ws.append(["e", "f"])

    assert ws.cell(row=1, column=1).value == "a"
    assert ws.cell(row=1, column=2).value == "b"
    assert ws.cell(row=2, column=1).value == "c"
    assert ws.cell(row=2, column=2).value == "d"
    assert ws.cell(row=3, column=1).value == "e"
    assert ws.cell(row=3, column=2).value == "f"


def test_append_rows_batch():
    """ws.append_rows([[1,2],[3,4]]) stores cells correctly."""
    wb = Workbook()
    ws = wb.active
    ws.append_rows([[1, 2], [3, 4]])

    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=1, column=2).value == 2
    assert ws.cell(row=2, column=1).value == 3
    assert ws.cell(row=2, column=2).value == 4


def test_append_rows_saves_correctly():
    """Batch append + save + load back with openpyxl verifies data integrity."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append_rows(
        [
            [1, "hello", True],
            [2, "world", False],
            [3.14, "pi", None],
        ]
    )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        wb.save(path)
        assert os.path.exists(path)

        # Load back with the real openpyxl to verify
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Data"]

        assert ws2.cell(row=1, column=1).value == 1
        assert ws2.cell(row=1, column=2).value == "hello"
        assert ws2.cell(row=1, column=3).value is True
        assert ws2.cell(row=2, column=1).value == 2
        assert ws2.cell(row=2, column=2).value == "world"
        assert ws2.cell(row=2, column=3).value is False
        assert abs(ws2.cell(row=3, column=1).value - 3.14) < 1e-10
        assert ws2.cell(row=3, column=2).value == "pi"
        assert ws2.cell(row=3, column=3).value is None
    finally:
        os.unlink(path)


def test_append_mixed_types():
    """Strings, numbers, booleans, None, and formulas in batch."""
    wb = Workbook()
    ws = wb.active
    ws.append_rows(
        [
            ["text", 42, 3.14, True, False, None, "=SUM(B1:C1)"],
        ]
    )

    assert ws.cell(row=1, column=1).value == "text"
    assert ws.cell(row=1, column=2).value == 42
    assert ws.cell(row=1, column=3).value == 3.14
    assert ws.cell(row=1, column=4).value is True
    assert ws.cell(row=1, column=5).value is False
    assert ws.cell(row=1, column=6).value is None
    assert ws.cell(row=1, column=7).value == "=SUM(B1:C1)"


def test_append_after_cell():
    """Use ws.cell() first, then ws.append() -- rows should not overlap."""
    wb = Workbook()
    ws = wb.active

    # Write to row 1 and row 3 using cell()
    ws.cell(row=1, column=1, value="header1")
    ws.cell(row=1, column=2, value="header2")
    ws.cell(row=3, column=1, value="row3")

    # append() should go to row 4 (after the max existing row 3)
    ws.append(["appended1", "appended2"])

    assert ws.cell(row=1, column=1).value == "header1"
    assert ws.cell(row=1, column=2).value == "header2"
    assert ws.cell(row=3, column=1).value == "row3"
    assert ws.cell(row=4, column=1).value == "appended1"
    assert ws.cell(row=4, column=2).value == "appended2"


def test_append_rows_after_cell():
    """Use ws.cell() first, then ws.append_rows() -- rows should not overlap."""
    wb = Workbook()
    ws = wb.active

    ws.cell(row=2, column=1, value="existing")

    ws.append_rows([[10, 20], [30, 40]])

    assert ws.cell(row=2, column=1).value == "existing"
    assert ws.cell(row=3, column=1).value == 10
    assert ws.cell(row=3, column=2).value == 20
    assert ws.cell(row=4, column=1).value == 30
    assert ws.cell(row=4, column=2).value == 40


def test_append_then_append_rows():
    """Mix append() and append_rows() -- they continue sequentially."""
    wb = Workbook()
    ws = wb.active

    ws.append(["row1_a", "row1_b"])
    ws.append(["row2_a", "row2_b"])
    ws.append_rows([["row3_a", "row3_b"], ["row4_a", "row4_b"]])

    assert ws.cell(row=1, column=1).value == "row1_a"
    assert ws.cell(row=2, column=1).value == "row2_a"
    assert ws.cell(row=3, column=1).value == "row3_a"
    assert ws.cell(row=4, column=1).value == "row4_a"


def test_append_rows_batch_saves_with_formulas():
    """Formulas via batch append should survive save/load roundtrip."""
    wb = Workbook()
    ws = wb.active
    ws.append_rows(
        [
            [10, 20],
            [30, 40],
            ["=SUM(A1:A2)", "=SUM(B1:B2)"],
        ]
    )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active

        assert ws2.cell(row=1, column=1).value == 10
        assert ws2.cell(row=1, column=2).value == 20
        assert ws2.cell(row=2, column=1).value == 30
        assert ws2.cell(row=2, column=2).value == 40
        # Formulas are stored as strings starting with =
        assert ws2.cell(row=3, column=1).value == "=SUM(A1:A2)"
        assert ws2.cell(row=3, column=2).value == "=SUM(B1:B2)"
    finally:
        os.unlink(path)
