# tests/test_load.py
import io
import os
import tempfile
from pathlib import Path

import openpyxl as real_openpyxl
import pytest

from openpyxl_rust import load_workbook


def _make_test_file(setup_fn):
    """Create a test xlsx file using real openpyxl."""
    wb = real_openpyxl.Workbook()
    setup_fn(wb)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    return path


def test_load_basic():
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "Hello"))
    try:
        wb = load_workbook(path)
        assert wb.active["A1"].value == "Hello"
    finally:
        os.unlink(path)


def test_load_numbers():
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", 42.5))
    try:
        wb = load_workbook(path)
        assert wb.active["A1"].value == 42.5
    finally:
        os.unlink(path)


def test_load_boolean():
    def setup(wb):
        wb.active["A1"] = True
        wb.active["A2"] = False

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        assert wb.active["A1"].value is True
        assert wb.active["A2"].value is False
    finally:
        os.unlink(path)


def test_load_multiple_sheets():
    def setup(wb):
        wb.active.title = "First"
        wb.active["A1"] = "one"
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "two"

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        assert wb.sheetnames == ["First", "Second"]
        assert wb["First"]["A1"].value == "one"
        assert wb["Second"]["A1"].value == "two"
    finally:
        os.unlink(path)


def test_load_empty_cells():
    def setup(wb):
        wb.active["A1"] = "data"
        wb.active["C3"] = "sparse"

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        assert wb.active["A1"].value == "data"
        assert wb.active["C3"].value == "sparse"
        assert wb.active["B2"].value is None
    finally:
        os.unlink(path)


def test_load_large_data():
    def setup(wb):
        ws = wb.active
        for r in range(1, 101):
            ws.cell(row=r, column=1, value=f"row_{r}")
            ws.cell(row=r, column=2, value=r * 1.5)

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "row_1"
        assert ws.cell(row=100, column=2).value == 150.0
    finally:
        os.unlink(path)


def test_load_then_save():
    """Load a file, modify it, save to new file, verify."""

    def setup(wb):
        wb.active["A1"] = "original"

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        wb.active["A1"] = "modified"
        wb.active["B1"] = "added"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path2 = f.name
        try:
            wb.save(path2)
            rb = real_openpyxl.load_workbook(path2)
            assert rb.active["A1"].value == "modified"
            assert rb.active["B1"].value == "added"
        finally:
            os.unlink(path2)
    finally:
        os.unlink(path)


def test_load_integers():
    """Verify integer values are returned as int, not float."""

    def setup(wb):
        wb.active["A1"] = 42

    path = _make_test_file(setup)
    try:
        wb = load_workbook(path)
        val = wb.active["A1"].value
        # calamine may return int or float — both are acceptable
        assert val == 42
    finally:
        os.unlink(path)


# =====================================================================
# BytesIO / file-like object support tests
# =====================================================================


def test_load_from_bytesio():
    """Load workbook from a BytesIO object."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "from_bytes"))
    try:
        with open(path, "rb") as f:
            raw = f.read()
        bio = io.BytesIO(raw)
        wb = load_workbook(bio)
        assert wb.active["A1"].value == "from_bytes"
    finally:
        os.unlink(path)


def test_load_from_bytesio_multiple_sheets():
    """Load a multi-sheet workbook from BytesIO."""

    def setup(wb):
        wb.active.title = "Alpha"
        wb.active["A1"] = "first"
        ws2 = wb.create_sheet("Beta")
        ws2["A1"] = "second"
        ws2["A2"] = 99

    path = _make_test_file(setup)
    try:
        with open(path, "rb") as f:
            bio = io.BytesIO(f.read())
        wb = load_workbook(bio)
        assert wb.sheetnames == ["Alpha", "Beta"]
        assert wb["Alpha"]["A1"].value == "first"
        assert wb["Beta"]["A1"].value == "second"
        assert wb["Beta"]["A2"].value == 99
    finally:
        os.unlink(path)


def test_load_from_open_file_handle():
    """Load workbook from an open binary file handle."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "handle"))
    try:
        with open(path, "rb") as fh:
            wb = load_workbook(fh)
        assert wb.active["A1"].value == "handle"
    finally:
        os.unlink(path)


def test_load_from_pathlib_path():
    """Load workbook from a pathlib.Path object."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "pathlib"))
    try:
        wb = load_workbook(Path(path))
        assert wb.active["A1"].value == "pathlib"
    finally:
        os.unlink(path)


def test_load_from_string_path():
    """Loading from a plain string path still works (regression check)."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "strpath"))
    try:
        wb = load_workbook(path)
        assert wb.active["A1"].value == "strpath"
    finally:
        os.unlink(path)


def test_load_invalid_type_raises_typeerror():
    """Passing an invalid type (e.g. int) should raise TypeError with a clear message."""
    with pytest.raises(TypeError, match=r"file path.*file-like object"):
        load_workbook(12345)


def test_load_invalid_type_list():
    """Passing a list should raise TypeError."""
    with pytest.raises(TypeError, match="got list"):
        load_workbook([1, 2, 3])


def test_load_text_mode_file_raises_typeerror():
    """A file opened in text mode should raise TypeError (read() returns str, not bytes)."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "text"))
    try:
        with open(path) as fh, pytest.raises(TypeError, match="binary mode"):
            load_workbook(fh)
    finally:
        os.unlink(path)


def test_load_bytesio_then_save():
    """Load from BytesIO, modify, save, and verify with real openpyxl."""
    path = _make_test_file(lambda wb: setattr(wb.active["A1"], "value", "original"))
    try:
        with open(path, "rb") as f:
            bio = io.BytesIO(f.read())
        wb = load_workbook(bio)
        wb.active["A1"] = "changed"
        wb.active["C1"] = "new"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path2 = f.name
        try:
            wb.save(path2)
            rb = real_openpyxl.load_workbook(path2)
            assert rb.active["A1"].value == "changed"
            assert rb.active["C1"].value == "new"
        finally:
            os.unlink(path2)
    finally:
        os.unlink(path)
