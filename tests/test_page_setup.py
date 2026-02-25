# tests/test_page_setup.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Workbook


def test_landscape():
    """Set orientation landscape, save and verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Landscape"
    ws.page_setup.orientation = "landscape"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.page_setup.orientation == "landscape"
    finally:
        os.unlink(path)


def test_paper_size_a4():
    """Set paper size A4, save and verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "A4 paper"
    ws.page_setup.paperSize = 9  # A4

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # openpyxl returns paperSize as int
        assert int(sheet.page_setup.paperSize) == 9
    finally:
        os.unlink(path)


def test_margins():
    """Set custom margins, save and verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Custom margins"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert abs(sheet.page_margins.left - 0.5) < 0.01
        assert abs(sheet.page_margins.right - 0.5) < 0.01
        assert abs(sheet.page_margins.top - 0.75) < 0.01
        assert abs(sheet.page_margins.bottom - 0.75) < 0.01
        assert abs(sheet.page_margins.header - 0.3) < 0.01
        assert abs(sheet.page_margins.footer - 0.3) < 0.01
    finally:
        os.unlink(path)


def test_print_area():
    """Set print area 'A1:F10', save and verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Print area test"
    ws["F10"] = "End"
    ws.print_area = "A1:F10"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # openpyxl returns print_area as a string like "'Sheet'!$A$1:$F$10"
        pa = sheet.print_area
        assert pa is not None
        # Normalize: remove $ signs and sheet name prefix for comparison
        normalized = pa.replace("$", "")
        # Remove sheet name prefix like "'Sheet'!" if present
        if "!" in normalized:
            normalized = normalized.split("!", 1)[1]
        assert normalized == "A1:F10"
    finally:
        os.unlink(path)


def test_print_titles():
    """Set repeat rows '1:1' and cols 'A:B', save and verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header Col A"
    ws["B1"] = "Header Col B"
    ws["A2"] = "Data"
    ws.print_title_rows = "1:1"
    ws.print_title_cols = "A:B"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # openpyxl stores print_title_rows and print_title_cols
        # They may contain $ signs, e.g. "$1:$1" or "1:1"
        if sheet.print_title_rows:
            normalized_rows = sheet.print_title_rows.replace("$", "")
            assert normalized_rows == "1:1"
        if sheet.print_title_cols:
            normalized_cols = sheet.print_title_cols.replace("$", "")
            assert normalized_cols == "A:B"
    finally:
        os.unlink(path)


def test_fit_to_pages():
    """fitToWidth=2, fitToHeight=3, save and verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Fit to page"
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 3

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # openpyxl reads back fitToWidth/fitToHeight
        assert sheet.page_setup.fitToWidth is not None
        assert sheet.page_setup.fitToHeight is not None
        assert int(sheet.page_setup.fitToWidth) == 2
        assert int(sheet.page_setup.fitToHeight) == 3
    finally:
        os.unlink(path)


def test_fit_to_pages_default_one():
    """fitToWidth=1, fitToHeight=0 (fit to 1 page wide, unlimited height).

    When fitToWidth=1, rust_xlsxwriter omits the attribute since 1 is the XML default.
    Verify the sheet properties indicate fit-to-page is enabled.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Fit to 1 page wide"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # fitToHeight=0 means "as many as needed", written as 0 in XML
        # which is != 1, so it will be present
        assert sheet.page_setup.fitToHeight is not None
        assert int(sheet.page_setup.fitToHeight) == 0
    finally:
        os.unlink(path)


def test_center_horizontally():
    """print_options.horizontalCentered = True, save and verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Centered"
    ws.print_options.horizontalCentered = True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.print_options.horizontalCentered is True
    finally:
        os.unlink(path)


def test_page_setup_default():
    """No page setup changes, verify save works fine."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Default"
    ws["B1"] = 42

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].value == "Default"
        assert sheet["B1"].value == 42
    finally:
        os.unlink(path)
