import openpyxl

from openpyxl_rust import DataValidation, Workbook


def test_dropdown_list(tmp_path):
    """DataValidation type='list' with inline items, verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Pick one:"

    dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_list.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    # Verify with openpyxl
    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.type == "list"
    # The formula should contain the list items
    assert "Dog" in dv2.formula1
    assert "Cat" in dv2.formula1
    assert "Bat" in dv2.formula1
    wb2.close()


def test_whole_number_greater(tmp_path):
    """DataValidation type='whole', operator='greaterThan', verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Enter positive integer:"

    dv = DataValidation(type="whole", operator="greaterThan", formula1="0")
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_whole.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.type == "whole"
    assert dv2.operator == "greaterThan"
    wb2.close()


def test_decimal_between(tmp_path):
    """DataValidation type='decimal', operator='between', formula1=0, formula2=1."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Enter decimal 0-1:"

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_decimal.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.type == "decimal"
    # "between" is the default operator in Excel XML, so openpyxl may report it as None
    assert dv2.operator in ("between", None)
    assert dv2.formula1 == "0"
    assert dv2.formula2 == "1"
    wb2.close()


def test_text_length(tmp_path):
    """DataValidation type='textLength', operator='lessThanOrEqual', formula1=15."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Max 15 chars:"

    dv = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="15")
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_textlen.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.type == "textLength"
    assert dv2.operator == "lessThanOrEqual"
    wb2.close()


def test_error_messages(tmp_path):
    """Set error title + message + style, verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Enter value:"

    dv = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="100",
        showErrorMessage=True,
        errorTitle="Bad Value",
        error="Please enter 1-100",
        errorStyle="warning",
    )
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_error.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.errorTitle == "Bad Value"
    assert dv2.error == "Please enter 1-100"
    assert dv2.errorStyle == "warning"
    wb2.close()


def test_input_messages(tmp_path):
    """Set prompt title + message, verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Enter value:"

    dv = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="10",
        showInputMessage=True,
        promptTitle="Rating",
        prompt="Enter a rating 1-10",
    )
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_input.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.promptTitle == "Rating"
    assert dv2.prompt == "Enter a rating 1-10"
    wb2.close()


def test_validation_save_valid(tmp_path):
    """Save + load, verify file valid with multiple validations."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "List:"
    ws["A2"] = "Number:"
    ws["A3"] = "Text:"

    dv1 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv1.add("B1")

    dv2 = DataValidation(type="whole", operator="greaterThan", formula1="0")
    dv2.add("B2")

    dv3 = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="50")
    dv3.add("B3")

    ws.add_data_validation(dv1)
    ws.add_data_validation(dv2)
    ws.add_data_validation(dv3)

    out = tmp_path / "dv_multi.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert ws2["A1"].value == "List:"
    assert ws2["A2"].value == "Number:"
    assert ws2["A3"].value == "Text:"
    assert len(ws2.data_validations.dataValidation) == 3
    wb2.close()


def test_sqref_range(tmp_path):
    """DataValidation applied via sqref (range string) instead of add()."""
    wb = Workbook()
    ws = wb.active

    dv = DataValidation(type="list", formula1='"A,B,C"', sqref="A1:A10")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_sqref.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    wb2.close()


def test_custom_validation(tmp_path):
    """DataValidation type='custom' with a formula."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Uppercase only:"

    dv = DataValidation(type="custom", formula1="=AND(ISTEXT(B1),EXACT(B1,UPPER(B1)))")
    dv.add("B1")
    ws.add_data_validation(dv)

    out = tmp_path / "dv_custom.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert len(ws2.data_validations.dataValidation) == 1
    dv2 = ws2.data_validations.dataValidation[0]
    assert dv2.type == "custom"
    wb2.close()
