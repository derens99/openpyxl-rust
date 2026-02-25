# tests/test_hyperlinks.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Workbook


def test_hyperlink_string():
    """Set cell.hyperlink = URL string, save, verify file is valid."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Example"
    ws["A1"].hyperlink = "https://example.com"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        # Verify with real openpyxl
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].hyperlink is not None
        assert "example.com" in sheet["A1"].hyperlink.target
    finally:
        os.unlink(path)


def test_hyperlink_internal():
    """cell.hyperlink = '#Sheet2!A1' creates an internal link."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Target"

    ws1["A1"] = "Go to Sheet2"
    ws1["A1"].hyperlink = "#Sheet2!A1"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        # Verify with real openpyxl
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded["Sheet1"]
        assert sheet["A1"].hyperlink is not None
        target = sheet["A1"].hyperlink.target or ""
        location = sheet["A1"].hyperlink.location or ""
        # Internal links may appear in target or location depending on the library
        assert "Sheet2" in target or "Sheet2" in location
    finally:
        os.unlink(path)


def test_hyperlink_with_value():
    """Set cell value and hyperlink, verify both survive save."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Click me"
    ws["A1"].hyperlink = "https://example.com"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # The cell should have a hyperlink
        assert sheet["A1"].hyperlink is not None
        assert "example.com" in sheet["A1"].hyperlink.target
        # The cell should also have a value (either the original text or the URL)
        assert sheet["A1"].value is not None
    finally:
        os.unlink(path)


def test_hyperlink_multiple():
    """Multiple cells with hyperlinks."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Example"
    ws["A1"].hyperlink = "https://example.com"
    ws["A2"] = "Python"
    ws["A2"].hyperlink = "https://python.org"
    ws["A3"] = "Rust"
    ws["A3"].hyperlink = "https://rust-lang.org"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].hyperlink is not None
        assert "example.com" in sheet["A1"].hyperlink.target
        assert sheet["A2"].hyperlink is not None
        assert "python.org" in sheet["A2"].hyperlink.target
        assert sheet["A3"].hyperlink is not None
        assert "rust-lang.org" in sheet["A3"].hyperlink.target
    finally:
        os.unlink(path)


def test_hyperlink_save_valid():
    """Save with hyperlinks, load with openpyxl, verify the hyperlink target is present."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Link"
    ws["A1"].hyperlink = "https://www.example.com/page?q=1"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        link = sheet["A1"].hyperlink
        assert link is not None
        assert link.target == "https://www.example.com/page?q=1"
    finally:
        os.unlink(path)
