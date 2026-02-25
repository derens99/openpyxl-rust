# tests/test_cell_style_compat.py
"""Tests for cell.data_type property and PatternFill fgColor/bgColor aliases."""
import os
import tempfile
from datetime import datetime, date, time

from openpyxl_rust import Workbook
from openpyxl_rust.cell import Cell
from openpyxl_rust.styles import PatternFill


# ── data_type property tests ────────────────────────────────────────

def test_data_type_string():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    assert ws["A1"].data_type == "s"


def test_data_type_number_int():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 42
    assert ws["A1"].data_type == "n"


def test_data_type_number_float():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 3.14
    assert ws["A1"].data_type == "n"


def test_data_type_bool():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = True
    assert ws["A1"].data_type == "b"


def test_data_type_date():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = date(2026, 2, 24)
    assert ws["A1"].data_type == "d"


def test_data_type_datetime():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = datetime(2026, 2, 24, 12, 30, 0)
    assert ws["A1"].data_type == "d"


def test_data_type_time():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = time(14, 30, 0)
    assert ws["A1"].data_type == "d"


def test_data_type_formula():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=SUM(A1:A5)"
    assert ws["A1"].data_type == "f"


def test_data_type_none():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = None
    assert ws["A1"].data_type == "n"


# ── TYPE_* class constants ──────────────────────────────────────────

def test_data_type_constants():
    assert Cell.TYPE_STRING == "s"
    assert Cell.TYPE_FORMULA == "f"
    assert Cell.TYPE_NUMERIC == "n"
    assert Cell.TYPE_BOOL == "b"
    assert Cell.TYPE_NULL == "n"
    assert Cell.TYPE_INLINE == "s"
    assert Cell.TYPE_ERROR == "e"
    assert Cell.TYPE_FORMULA_CACHE_STRING == "s"


# ── PatternFill fgColor / bgColor aliases ───────────────────────────

def test_fgcolor_constructor():
    fill = PatternFill(fgColor="FF0000")
    assert fill.start_color == "FF0000"


def test_bgcolor_constructor():
    fill = PatternFill(bgColor="00FF00")
    assert fill.end_color == "00FF00"


def test_fgcolor_property_getter():
    fill = PatternFill(start_color="AABBCC")
    assert fill.fgColor == "AABBCC"


def test_fgcolor_property_setter():
    fill = PatternFill()
    fill.fgColor = "112233"
    assert fill.start_color == "112233"
    assert fill.fgColor == "112233"


def test_bgcolor_property_getter():
    fill = PatternFill(end_color="DDEEFF")
    assert fill.bgColor == "DDEEFF"


def test_bgcolor_property_setter():
    fill = PatternFill()
    fill.bgColor = "445566"
    assert fill.end_color == "445566"
    assert fill.bgColor == "445566"


def test_fgcolor_save_roundtrip():
    """Save with fgColor alias, read back with openpyxl, verify color applied."""
    import openpyxl as real_openpyxl
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Red"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        cell = rb.active["A1"]
        assert cell.fill.patternType == "solid"
        assert "FF0000" in cell.fill.start_color.rgb
    finally:
        os.unlink(path)
