"""Tests for conditional formatting support."""
import os
import tempfile
import pytest

from openpyxl_rust import Workbook
from openpyxl_rust.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, CellIsRule, FormulaRule
)
from openpyxl_rust.styles.fills import PatternFill
from openpyxl_rust.styles.fonts import Font
from openpyxl_rust.styles.borders import Border, Side


@pytest.fixture
def wb_with_data():
    """Create a workbook with sample numeric data in A1:A10."""
    wb = Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i * 10)
    return wb, ws


def _save_and_reload(wb):
    """Save a workbook and reload it with openpyxl to verify structure."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        import openpyxl
        wb2 = openpyxl.load_workbook(path)
        return wb2
    finally:
        if os.path.exists(path):
            os.unlink(path)


class TestColorScale:
    def test_color_scale_2(self, wb_with_data):
        """Two-color scale (min red -> max green)."""
        wb, ws = wb_with_data
        rule = ColorScaleRule(
            start_type="min", start_color="FF0000",
            end_type="max", end_color="00FF00"
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        # There should be at least one conditional format
        assert len(list(cf_rules)) > 0

    def test_color_scale_3(self, wb_with_data):
        """Three-color scale with midpoint."""
        wb, ws = wb_with_data
        rule = ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00FF00"
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0


class TestDataBar:
    def test_data_bar(self, wb_with_data):
        """Basic data bar with custom color."""
        wb, ws = wb_with_data
        rule = DataBarRule(
            start_type="min", end_type="max",
            color="638EC6"
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_data_bar_bar_only(self, wb_with_data):
        """Data bar with showValue=False (bar only, no numbers)."""
        wb, ws = wb_with_data
        rule = DataBarRule(
            start_type="min", end_type="max",
            color="638EC6", showValue=False
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0


class TestIconSet:
    def test_icon_set(self, wb_with_data):
        """3Arrows icon set."""
        wb, ws = wb_with_data
        rule = IconSetRule(icon_style="3Arrows")
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_icon_set_reverse(self, wb_with_data):
        """Icon set with reverse."""
        wb, ws = wb_with_data
        rule = IconSetRule(icon_style="3Arrows", reverse=True)
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0


class TestCellIs:
    def test_cell_is_less_than(self, wb_with_data):
        """CellIsRule with lessThan operator and fill format."""
        wb, ws = wb_with_data
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")
        rule = CellIsRule(
            operator="lessThan",
            formula=["50"],
            fill=red_fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_cell_is_between(self, wb_with_data):
        """CellIsRule with between operator."""
        wb, ws = wb_with_data
        green_fill = PatternFill(fill_type="solid", start_color="C6EFCE")
        rule = CellIsRule(
            operator="between",
            formula=["30", "70"],
            fill=green_fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_cell_is_with_font(self, wb_with_data):
        """CellIsRule with font formatting."""
        wb, ws = wb_with_data
        red_font = Font(bold=True, color="9C0006")
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")
        rule = CellIsRule(
            operator="greaterThan",
            formula=["80"],
            font=red_font,
            fill=red_fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_cell_is_equal(self, wb_with_data):
        """CellIsRule with equal operator."""
        wb, ws = wb_with_data
        fill = PatternFill(fill_type="solid", start_color="FFFF00")
        rule = CellIsRule(
            operator="equal",
            formula=["50"],
            fill=fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0


class TestFormulaRule:
    def test_formula_rule(self, wb_with_data):
        """FormulaRule with formula and fill."""
        wb, ws = wb_with_data
        fill = PatternFill(fill_type="solid", start_color="FFFF00")
        rule = FormulaRule(
            formula=["ISBLANK(A1)"],
            fill=fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0

    def test_formula_rule_with_font(self, wb_with_data):
        """FormulaRule with font and fill."""
        wb, ws = wb_with_data
        font = Font(bold=True, color="FF0000")
        fill = PatternFill(fill_type="solid", start_color="FFCCCC")
        rule = FormulaRule(
            formula=["MOD(ROW(),2)=0"],
            font=font,
            fill=fill
        )
        ws.conditional_formatting.add("A1:A10", rule)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = ws2.conditional_formatting
        assert len(list(cf_rules)) > 0


class TestMultipleRules:
    def test_multiple_rules(self, wb_with_data):
        """Multiple rules on the same range."""
        wb, ws = wb_with_data
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")
        green_fill = PatternFill(fill_type="solid", start_color="C6EFCE")

        rule1 = CellIsRule(
            operator="lessThan",
            formula=["50"],
            fill=red_fill
        )
        rule2 = CellIsRule(
            operator="greaterThanOrEqual",
            formula=["50"],
            fill=green_fill
        )
        ws.conditional_formatting.add("A1:A10", rule1)
        ws.conditional_formatting.add("A1:A10", rule2)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        # Should have conditional formatting applied
        cf_rules = list(ws2.conditional_formatting)
        assert len(cf_rules) >= 1

    def test_multiple_rules_different_ranges(self, wb_with_data):
        """Multiple rules on different ranges."""
        wb, ws = wb_with_data
        # Add data to column B too
        for i in range(1, 11):
            ws.cell(row=i, column=2, value=i * 5)

        rule1 = ColorScaleRule(
            start_type="min", start_color="FF0000",
            end_type="max", end_color="00FF00"
        )
        rule2 = DataBarRule(color="638EC6")

        ws.conditional_formatting.add("A1:A10", rule1)
        ws.conditional_formatting.add("B1:B10", rule2)

        wb2 = _save_and_reload(wb)
        ws2 = wb2.active
        cf_rules = list(ws2.conditional_formatting)
        assert len(cf_rules) >= 2


class TestConditionalFormatSaveValid:
    def test_conditional_format_save_valid(self, wb_with_data):
        """Save + load, verify file opens without errors using openpyxl."""
        wb, ws = wb_with_data

        # Apply one of each type of conditional format
        rule1 = ColorScaleRule(
            start_type="min", start_color="FF0000",
            end_type="max", end_color="00FF00"
        )
        ws.conditional_formatting.add("A1:A10", rule1)

        rule2 = DataBarRule(color="638EC6")
        ws.conditional_formatting.add("A1:A10", rule2)

        rule3 = IconSetRule(icon_style="3Arrows")
        ws.conditional_formatting.add("A1:A10", rule3)

        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")
        rule4 = CellIsRule(operator="lessThan", formula=["50"], fill=red_fill)
        ws.conditional_formatting.add("A1:A10", rule4)

        fill = PatternFill(fill_type="solid", start_color="FFFF00")
        rule5 = FormulaRule(formula=["ISBLANK(A1)"], fill=fill)
        ws.conditional_formatting.add("A1:A10", rule5)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            wb.save(path)
            # Verify file exists and is non-empty
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0

            # Verify openpyxl can read it without errors
            import openpyxl
            wb2 = openpyxl.load_workbook(path)
            ws2 = wb2.active

            # There should be conditional formatting on the sheet
            cf_rules = list(ws2.conditional_formatting)
            assert len(cf_rules) >= 1

            # Values should still be intact
            assert ws2["A1"].value == 10
            assert ws2["A10"].value == 100
        finally:
            if os.path.exists(path):
                os.unlink(path)
