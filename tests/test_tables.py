"""Tests for Table support."""

import openpyxl

from openpyxl_rust import Table, TableColumn, TableStyleInfo, Workbook


class TestTableCreation:
    def test_basic_table(self, tmp_path):
        """Create a basic table and verify it saves to valid xlsx."""
        wb = Workbook()
        ws = wb.active
        # Write header + data
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])
        ws.append(["Charlie", 35, "CHI"])

        tab = Table(displayName="People", ref="A1:C4")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
        ws.add_table(tab)

        path = str(tmp_path / "table_basic.xlsx")
        wb.save(path)

        # Verify with openpyxl
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        assert len(ws2.tables) == 1
        t = next(iter(ws2.tables.values()))
        assert t.displayName == "People"
        assert t.ref == "A1:C4"

    def test_table_style_light(self, tmp_path):
        """Test with a Light table style."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Product", "Price"])
        ws.append(["Apple", 1.50])
        ws.append(["Banana", 0.75])

        tab = Table(displayName="Products", ref="A1:B3")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1")
        ws.add_table(tab)

        path = str(tmp_path / "table_light.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        t = next(iter(ws2.tables.values()))
        assert t.displayName == "Products"
        assert t.tableStyleInfo.name == "TableStyleLight1"

    def test_table_style_dark(self, tmp_path):
        """Test with a Dark table style."""
        wb = Workbook()
        ws = wb.active
        ws.append(["X", "Y"])
        ws.append([1, 2])

        tab = Table(displayName="DarkTable", ref="A1:B2")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleDark3")
        ws.add_table(tab)

        path = str(tmp_path / "table_dark.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        t = next(iter(wb2.active.tables.values()))
        assert t.tableStyleInfo.name == "TableStyleDark3"

    def test_table_with_columns(self, tmp_path):
        """Test table with explicit column definitions."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Col1", "Col2", "Col3"])
        ws.append([10, 20, 30])
        ws.append([40, 50, 60])

        tab = Table(displayName="WithCols", ref="A1:C3")
        tab.tableColumns = [
            TableColumn(id=1, name="Col1"),
            TableColumn(id=2, name="Col2"),
            TableColumn(id=3, name="Col3"),
        ]
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
        ws.add_table(tab)

        path = str(tmp_path / "table_cols.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        t = next(iter(wb2.active.tables.values()))
        assert t.displayName == "WithCols"
        cols = t.tableColumns
        assert len(cols) == 3
        assert cols[0].name == "Col1"
        assert cols[1].name == "Col2"
        assert cols[2].name == "Col3"

    def test_table_row_stripes(self, tmp_path):
        """Test table style with row stripes and column stripes toggles."""
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])

        tab = Table(displayName="Striped", ref="A1:B2")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=True,
            showFirstColumn=True,
            showLastColumn=True,
        )
        ws.add_table(tab)

        path = str(tmp_path / "table_stripes.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        t = next(iter(wb2.active.tables.values()))
        si = t.tableStyleInfo
        assert si.showRowStripes is True
        assert si.showColumnStripes is True
        assert si.showFirstColumn is True
        assert si.showLastColumn is True

    def test_multiple_tables_on_sheet(self, tmp_path):
        """Test adding multiple tables to one sheet."""
        wb = Workbook()
        ws = wb.active
        # Table 1 data
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws.append([3, 4])
        # Table 2 data (separate area)
        ws.cell(row=1, column=5, value="X")
        ws.cell(row=1, column=6, value="Y")
        ws.cell(row=2, column=5, value=10)
        ws.cell(row=2, column=6, value=20)

        tab1 = Table(displayName="Table1", ref="A1:B3")
        tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1")
        ws.add_table(tab1)

        tab2 = Table(displayName="Table2", ref="E1:F2")
        tab2.tableStyleInfo = TableStyleInfo(name="TableStyleLight5")
        ws.add_table(tab2)

        path = str(tmp_path / "multi_table.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        assert len(ws2.tables) == 2
        names = {t.displayName for t in ws2.tables.values()}
        assert "Table1" in names
        assert "Table2" in names

    def test_table_no_style(self, tmp_path):
        """Test table without explicit style info (uses rust_xlsxwriter defaults)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["H1", "H2"])
        ws.append([1, 2])

        tab = Table(displayName="NoStyle", ref="A1:B2")
        ws.add_table(tab)

        path = str(tmp_path / "table_nostyle.xlsx")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        t = next(iter(wb2.active.tables.values()))
        assert t.displayName == "NoStyle"
