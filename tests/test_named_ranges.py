# tests/test_named_ranges.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import DefinedName, Workbook


def test_define_name_basic():
    """Define 'Sales' = 'Sheet!$A$1:$A$10', save, verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=i * 10)

    dn = DefinedName("Sales", attr_text="Sheet!$A$1:$A$10")
    wb.defined_names.add(dn)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        assert "Sales" in loaded.defined_names
        assert "Sheet!$A$1:$A$10" in loaded.defined_names["Sales"].attr_text
    finally:
        os.unlink(path)


def test_define_name_constant():
    """Define 'Rate' = '0.96', save, verify file is valid."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Rate"

    dn = DefinedName("Rate", attr_text="0.96")
    wb.defined_names.add(dn)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        assert "Rate" in loaded.defined_names
    finally:
        os.unlink(path)


def test_define_name_multiple():
    """Define multiple named ranges, verify all present."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Sales"
    ws["B1"] = "Costs"
    ws["C1"] = "Profit"

    definitions = [
        ("SalesRange", "Sheet!$A$1:$A$100"),
        ("CostsRange", "Sheet!$B$1:$B$100"),
        ("ProfitRange", "Sheet!$C$1:$C$100"),
    ]
    for name, formula in definitions:
        dn = DefinedName(name, attr_text=formula)
        wb.defined_names.add(dn)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        assert "SalesRange" in loaded.defined_names
        assert "CostsRange" in loaded.defined_names
        assert "ProfitRange" in loaded.defined_names
    finally:
        os.unlink(path)


def test_define_name_save_valid():
    """Save + load, verify file valid and names accessible."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Data"
    for i in range(2, 12):
        ws.cell(row=i, column=1, value=i)

    dn = DefinedName("DataRange", attr_text="Sheet!$A$1:$A$11")
    wb.defined_names.add(dn)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        # Load with openpyxl and verify
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet["A1"].value == "Data"
        assert sheet["A2"].value == 2

        assert "DataRange" in loaded.defined_names
        assert "Sheet!$A$1:$A$11" in loaded.defined_names["DataRange"].attr_text
    finally:
        os.unlink(path)
