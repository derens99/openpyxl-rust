# tests/test_protection.py
import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Workbook


def test_protection_disabled_by_default():
    """Verify protection.sheet is False by default."""
    wb = Workbook()
    ws = wb.active
    assert ws.protection.sheet is False
    assert ws.protection._password is None


def test_protection_enable():
    """Enable protection via ws.protection.sheet = True, save and verify with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Protected"
    ws.protection.sheet = True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.protection.sheet is True
    finally:
        os.unlink(path)


def test_protection_password():
    """Set password via set_password(), verify sheet is protected with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Secret"
    ws.protection.set_password("secret")

    # set_password should also enable protection
    assert ws.protection.sheet is True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.protection.sheet is True
        # openpyxl stores the password hash, not the raw password
        # Just check that a password hash is set (non-empty)
        assert sheet.protection.password is not None
        assert len(sheet.protection.password) > 0
    finally:
        os.unlink(path)


def test_protection_custom_options():
    """Allow specific actions (format_cells=False means users CAN format), verify."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Custom"
    ws.protection.sheet = True
    # In openpyxl convention: True = protected (can't do), False = allowed (can do)
    # Setting format_cells=False means users CAN format cells
    ws.protection.format_cells = False
    ws.protection.insert_rows = False
    ws.protection.sort = False

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.protection.sheet is True
        # openpyxl reads back the same convention:
        # formatCells=False means user CAN format cells
        assert sheet.protection.formatCells is False
        assert sheet.protection.insertRows is False
        assert sheet.protection.sort is False
        # These should remain at their defaults (True = protected)
        assert sheet.protection.deleteRows is True
        assert sheet.protection.deleteColumns is True
    finally:
        os.unlink(path)


def test_protection_save_valid():
    """Save + load, verify sheet is protected and file is valid."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Data"
    ws["B1"] = 42
    ws.protection.sheet = True

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

        # Load with openpyxl to verify
        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        assert sheet.protection.sheet is True
        # Data should still be readable
        assert sheet["A1"].value == "Data"
        assert sheet["B1"].value == 42
    finally:
        os.unlink(path)


def test_protection_enable_disable():
    """Test enable() and disable() methods."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Toggle"

    ws.protection.enable()
    assert ws.protection.sheet is True

    ws.protection.disable()
    assert ws.protection.sheet is False

    # Should save without protection when disabled
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)

        loaded = real_openpyxl.load_workbook(path)
        sheet = loaded.active
        # Protection should not be enabled
        assert sheet.protection.sheet is False
    finally:
        os.unlink(path)


def test_protection_password_property():
    """Test the password property getter/setter."""
    wb = Workbook()
    ws = wb.active

    assert ws.protection.password is None
    ws.protection.password = "mypassword"
    assert ws.protection.password == "mypassword"
