import os
import tempfile

import openpyxl as real_openpyxl

from openpyxl_rust import Workbook


def _save_and_reload(wb):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    wb.save(path)
    try:
        rb = real_openpyxl.load_workbook(path)
        return rb, path
    except Exception:
        os.unlink(path)
        raise


def test_empty_workbook():
    """Save a workbook with no data at all."""
    wb = Workbook()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        rb = real_openpyxl.load_workbook(path)
        assert len(rb.sheetnames) == 1
    finally:
        os.unlink(path)


def test_unicode_emoji():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello 🌍🚀"
    ws["A2"] = "日本語テスト"
    rb, path = _save_and_reload(wb)
    try:
        assert rb.active["A1"].value == "Hello 🌍🚀"
        assert rb.active["A2"].value == "日本語テスト"
    finally:
        os.unlink(path)


def test_unicode_cjk():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "中文字符"
    ws["A2"] = "한국어"
    ws["A3"] = "العربية"
    rb, path = _save_and_reload(wb)
    try:
        assert rb.active["A1"].value == "中文字符"
        assert rb.active["A2"].value == "한국어"
        assert rb.active["A3"].value == "العربية"
    finally:
        os.unlink(path)


def test_extreme_float():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1e308
    ws["A2"] = 1e-308
    ws["A3"] = -1e308
    rb, path = _save_and_reload(wb)
    try:
        assert abs(rb.active["A1"].value - 1e308) / 1e308 < 1e-10
        assert rb.active["A2"].value == 1e-308
    finally:
        os.unlink(path)


def test_very_long_string():
    wb = Workbook()
    ws = wb.active
    long_str = "A" * 32767
    ws["A1"] = long_str
    rb, path = _save_and_reload(wb)
    try:
        assert len(rb.active["A1"].value) == 32767
    finally:
        os.unlink(path)


def test_cell_overwrite():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "first"
    ws["A1"] = "second"
    assert ws["A1"].value == "second"
    rb, path = _save_and_reload(wb)
    try:
        assert rb.active["A1"].value == "second"
    finally:
        os.unlink(path)


def test_mixed_case_cell_reference():
    wb = Workbook()
    ws = wb.active
    ws["a1"] = "lower"
    assert ws["A1"].value == "lower"


def test_sheet_name_special_chars():
    wb = Workbook()
    ws = wb.create_sheet("Data & Report (2024)")
    ws["A1"] = "test"
    rb, path = _save_and_reload(wb)
    try:
        assert "Data & Report (2024)" in rb.sheetnames
    finally:
        os.unlink(path)


def test_max_column_reference():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=16384, value="max col")
    assert ws.cell(row=1, column=16384).value == "max col"


def test_empty_string_value():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = ""
    assert ws["A1"].value == ""
    rb, path = _save_and_reload(wb)
    try:
        val = rb.active["A1"].value
        assert val == "" or val is None
    finally:
        os.unlink(path)


def test_zero_value():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 0
    ws["A2"] = 0.0
    assert ws["A1"].value == 0
    rb, path = _save_and_reload(wb)
    try:
        assert rb.active["A1"].value == 0
    finally:
        os.unlink(path)


def test_large_int():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 2**53
    rb, path = _save_and_reload(wb)
    try:
        assert rb.active["A1"].value == 2**53
    finally:
        os.unlink(path)
