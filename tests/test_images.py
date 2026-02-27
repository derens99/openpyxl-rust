import struct
import zlib

import pytest

from openpyxl_rust import Image, Workbook


def _make_mini_png():
    """Create a minimal valid 1x1 white PNG."""
    # PNG signature
    sig = b"\x89PNG\r\n\x1a\n"
    # IHDR chunk
    ihdr_data = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)  # 1x1, 8-bit RGB
    ihdr_crc = zlib.crc32(b"IHDR" + ihdr_data) & 0xFFFFFFFF
    ihdr = struct.pack(">I", 13) + b"IHDR" + ihdr_data + struct.pack(">I", ihdr_crc)
    # IDAT chunk
    raw = zlib.compress(b"\x00\xff\xff\xff")  # filter byte + RGB white
    idat_crc = zlib.crc32(b"IDAT" + raw) & 0xFFFFFFFF
    idat = struct.pack(">I", len(raw)) + b"IDAT" + raw + struct.pack(">I", idat_crc)
    # IEND chunk
    iend_crc = zlib.crc32(b"IEND") & 0xFFFFFFFF
    iend = struct.pack(">I", 0) + b"IEND" + struct.pack(">I", iend_crc)
    return sig + ihdr + idat + iend


def test_image_from_bytes(tmp_path):
    """Image created from bytes, add to sheet, save, verify file valid."""
    png_data = _make_mini_png()
    img = Image(png_data)
    assert img._data == png_data

    wb = Workbook()
    ws = wb.active
    ws.add_image(img, "A1")
    out = tmp_path / "img_bytes.xlsx"
    wb.save(str(out))
    assert out.exists()
    assert out.stat().st_size > 0


def test_image_from_file(tmp_path):
    """Write PNG to temp file, Image(path), add, save, verify."""
    png_data = _make_mini_png()
    png_path = tmp_path / "test.png"
    png_path.write_bytes(png_data)

    img = Image(str(png_path))
    assert img._data == png_data

    wb = Workbook()
    ws = wb.active
    ws.add_image(img, "B2")
    out = tmp_path / "img_file.xlsx"
    wb.save(str(out))
    assert out.exists()
    assert out.stat().st_size > 0


def test_image_with_anchor(tmp_path):
    """add_image(img, 'C5'), save, verify anchor was set."""
    png_data = _make_mini_png()
    img = Image(png_data)
    assert img.anchor is None

    wb = Workbook()
    ws = wb.active
    ws.add_image(img, "C5")
    assert img.anchor == "C5"

    out = tmp_path / "img_anchor.xlsx"
    wb.save(str(out))
    assert out.exists()
    assert out.stat().st_size > 0


def test_image_save_valid(tmp_path):
    """Save + load with openpyxl, verify file opens without error."""
    png_data = _make_mini_png()
    img = Image(png_data)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws.add_image(img, "D1")
    out = tmp_path / "img_valid.xlsx"
    wb.save(str(out))

    # Verify the file can be opened by openpyxl (real)
    import openpyxl

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert ws2["A1"].value == "Hello"
    wb2.close()


def test_image_multiple(tmp_path):
    """Multiple images on different cells."""
    png_data = _make_mini_png()

    wb = Workbook()
    ws = wb.active

    img1 = Image(png_data)
    img2 = Image(png_data)
    img3 = Image(png_data)

    ws.add_image(img1, "A1")
    ws.add_image(img2, "E5")
    ws.add_image(img3, "J10")

    assert len(ws._images) == 3

    out = tmp_path / "img_multi.xlsx"
    wb.save(str(out))
    assert out.exists()
    assert out.stat().st_size > 0

    # Verify the file can be opened by openpyxl (real)
    import openpyxl

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    # openpyxl should load images
    assert len(ws2._images) == 3
    wb2.close()


def test_image_type_error():
    """Image() with invalid type raises TypeError."""
    with pytest.raises(TypeError):
        Image(12345)


def test_image_from_bytearray():
    """Image created from bytearray."""
    png_data = _make_mini_png()
    ba = bytearray(png_data)
    img = Image(ba)
    assert img._data == png_data
    assert isinstance(img._data, bytes)
