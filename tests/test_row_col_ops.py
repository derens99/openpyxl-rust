import os
import tempfile

from openpyxl_rust import Workbook


def test_insert_rows_basic():
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2])
    ws.append([3, 4])
    ws.insert_rows(2, amount=1)
    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=3, column=1).value == 3  # shifted from row 2


def test_delete_rows_basic():
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["c", "d"])
    ws.append(["e", "f"])
    ws.delete_rows(2, amount=1)
    assert ws.cell(row=1, column=1).value == "a"
    assert ws.cell(row=2, column=1).value == "e"  # shifted from row 3


def test_insert_cols_basic():
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    ws.insert_cols(2, amount=1)
    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=1, column=3).value == 2  # shifted from col 2


def test_delete_cols_basic():
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    ws.delete_cols(2, amount=1)
    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=1, column=2).value == 3  # shifted from col 3


def test_insert_rows_saves_correctly():
    wb = Workbook()
    ws = wb.active
    ws.append(["header1", "header2"])
    ws.append(["data1", "data2"])
    ws.insert_rows(2, amount=1)
    ws.cell(row=2, column=1, value="inserted")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        import openpyxl

        rb = openpyxl.load_workbook(path)
        rws = rb.active
        assert rws.cell(row=1, column=1).value == "header1"
        assert rws.cell(row=2, column=1).value == "inserted"
        assert rws.cell(row=3, column=1).value == "data1"
    finally:
        os.unlink(path)


def test_delete_multiple_rows():
    wb = Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.append([f"row{i}"])
    ws.delete_rows(2, amount=2)
    assert ws.cell(row=1, column=1).value == "row1"
    assert ws.cell(row=2, column=1).value == "row4"
    assert ws.cell(row=3, column=1).value == "row5"


def test_insert_rows_updates_merged():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A2:B2")
    ws.insert_rows(1, amount=1)
    assert ("A3", "B3") in ws.merged_cell_ranges
