# tests/test_datetime.py
import os
import tempfile
from datetime import datetime, date

import openpyxl as real_openpyxl
from openpyxl_rust import Workbook


def test_save_datetime():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = datetime(2024, 3, 15, 10, 30, 0)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_save_date():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = date(2024, 3, 15)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_compat_datetime():
    wb = Workbook()
    ws = wb.active
    dt = datetime(2024, 3, 15, 10, 30, 0)
    ws["A1"] = dt
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        val = rb.active["A1"].value
        assert isinstance(val, datetime)
        assert val.year == 2024
        assert val.month == 3
        assert val.day == 15
        assert val.hour == 10
        assert val.minute == 30
    finally:
        os.unlink(path)


def test_compat_date():
    wb = Workbook()
    ws = wb.active
    d = date(2024, 3, 15)
    ws["A1"] = d
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        val = rb.active["A1"].value
        assert val.year == 2024
        assert val.month == 3
        assert val.day == 15
    finally:
        os.unlink(path)


def test_datetime_with_custom_format():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = datetime(2024, 12, 25, 0, 0, 0)
    ws["A1"].number_format = "dd/mm/yyyy"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        rb = real_openpyxl.load_workbook(path)
        assert rb.active["A1"].number_format == "dd/mm/yyyy"
    finally:
        os.unlink(path)
