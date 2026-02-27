from openpyxl_rust.workbook import Workbook


def test_workbook_has_active_sheet():
    wb = Workbook()
    assert wb.active is not None
    assert wb.active.title == "Sheet"


def test_workbook_create_sheet():
    wb = Workbook()
    ws = wb.create_sheet("Data")
    assert ws.title == "Data"
    assert len(wb.sheetnames) == 2


def test_workbook_sheetnames():
    wb = Workbook()
    wb.create_sheet("A")
    wb.create_sheet("B")
    assert wb.sheetnames == ["Sheet", "A", "B"]


def test_workbook_getitem():
    wb = Workbook()
    wb.create_sheet("Data")
    ws = wb["Data"]
    assert ws.title == "Data"


def test_workbook_remove_sheet():
    wb = Workbook()
    ws2 = wb.create_sheet("Second")
    wb.create_sheet("Third")
    wb.remove(ws2)
    assert wb.sheetnames == ["Sheet", "Third"]
    assert len(wb._sheets) == 2


def test_workbook_remove_first_sheet():
    wb = Workbook()
    ws1 = wb.active
    wb.create_sheet("Second")
    wb.remove(ws1)
    assert wb.sheetnames == ["Second"]
    assert wb.active.title == "Second"


def test_workbook_remove_and_save():
    import os
    import tempfile

    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = "keep"
    ws2 = wb.create_sheet("Remove")
    ws2["A1"] = "gone"
    wb.remove(ws2)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        import openpyxl

        rb = openpyxl.load_workbook(path)
        assert rb.sheetnames == ["Sheet"]
        assert rb.active["A1"].value == "keep"
    finally:
        os.unlink(path)


def test_workbook_remove_nonexistent():
    import pytest

    wb = Workbook()
    from openpyxl_rust.worksheet import Worksheet

    ws_other = Worksheet(title="Fake")
    with pytest.raises(ValueError):
        wb.remove(ws_other)


def test_workbook_remove_reindex():
    import os
    import tempfile

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1["A1"] = "one"
    ws2 = wb.create_sheet("Middle")
    ws2["A1"] = "two"
    ws3 = wb.create_sheet("Last")
    ws3["A1"] = "three"
    wb.remove(ws2)
    ws3["B1"] = "extra"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        import openpyxl

        rb = openpyxl.load_workbook(path)
        assert rb.sheetnames == ["First", "Last"]
        assert rb["Last"]["A1"].value == "three"
        assert rb["Last"]["B1"].value == "extra"
    finally:
        os.unlink(path)


# --- Feature 1: wb.active setter ---


def test_active_setter_by_worksheet():
    """Setting active to a specific Worksheet object works."""
    wb = Workbook()
    ws2 = wb.create_sheet("Second")
    ws3 = wb.create_sheet("Third")
    wb.active = ws3
    assert wb.active is ws3
    assert wb.active.title == "Third"
    wb.active = ws2
    assert wb.active is ws2


def test_active_setter_by_index():
    """Setting active by integer index works."""
    wb = Workbook()
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    wb.active = 2
    assert wb.active.title == "Third"
    wb.active = 0
    assert wb.active.title == "Sheet"
    # Out-of-range index raises IndexError
    import pytest

    with pytest.raises(IndexError):
        wb.active = 5
    with pytest.raises(IndexError):
        wb.active = -1


def test_active_after_remove():
    """Active sheet index adjusts correctly after sheet removal."""
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet("Second")
    ws3 = wb.create_sheet("Third")
    # Set active to Third (index 2)
    wb.active = ws3
    assert wb.active is ws3
    # Remove Second (index 1) — active should stay on Third (now index 1)
    wb.remove(ws2)
    assert wb.active is ws3
    assert wb.active.title == "Third"
    # Remove Third (the active) — active should clamp to last sheet
    wb.active = ws3
    wb.remove(ws3)
    assert wb.active is ws1
    assert wb.active.title == "Sheet"


# --- Feature 2: Workbook.__iter__ ---


def test_iter_workbook():
    """Iterating over a workbook yields its worksheets."""
    wb = Workbook()
    wb.create_sheet("Second")
    wb.create_sheet("Third")
    sheets = list(wb)
    assert len(sheets) == 3
    assert sheets[0].title == "Sheet"
    assert sheets[1].title == "Second"
    assert sheets[2].title == "Third"
    # Works in a for loop
    titles = [ws.title for ws in wb]
    assert titles == ["Sheet", "Second", "Third"]


# --- Feature 3: Workbook.__len__ ---


def test_len_workbook():
    """len(wb) returns the number of worksheets."""
    wb = Workbook()
    assert len(wb) == 1
    wb.create_sheet("Second")
    assert len(wb) == 2
    wb.create_sheet("Third")
    assert len(wb) == 3
    wb.remove(wb["Second"])
    assert len(wb) == 2


# --- Feature 4: Sheet title uniqueness ---


def test_duplicate_title_auto_rename():
    """create_sheet with an existing title auto-renames with a number suffix."""
    wb = Workbook()
    ws1 = wb.active  # "Sheet"
    ws2 = wb.create_sheet("Sheet")
    assert ws2.title == "Sheet1"
    assert ws2.title != ws1.title
    assert wb.sheetnames == ["Sheet", "Sheet1"]


def test_title_setter_rejects_duplicate():
    """Renaming a worksheet title to an existing name raises ValueError."""
    import pytest

    wb = Workbook()
    ws2 = wb.create_sheet("Second")
    with pytest.raises(ValueError, match="already exists"):
        ws2.title = "Sheet"


def test_create_sheet_unique_titles():
    """Creating many sheets with the same base name produces unique titles."""
    wb = Workbook()
    # Default sheet is "Sheet"
    wb.create_sheet("Sheet")
    wb.create_sheet("Sheet")
    wb.create_sheet("Sheet")
    assert wb.sheetnames == ["Sheet", "Sheet1", "Sheet2", "Sheet3"]
    # All titles unique
    assert len(set(wb.sheetnames)) == len(wb.sheetnames)
