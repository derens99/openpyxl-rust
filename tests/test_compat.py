# tests/test_compat.py
"""Read back files written by openpyxl_rust using openpyxl to verify correctness."""
import os
import tempfile

import openpyxl as real_openpyxl
from openpyxl_rust import Workbook
from openpyxl_rust.styles import Font


def _write_and_read(setup_fn):
    """Helper: create workbook, apply setup_fn, save, read back with openpyxl."""
    wb = Workbook()
    setup_fn(wb)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        wb.save(path)
        return real_openpyxl.load_workbook(path)
    finally:
        os.unlink(path)


def test_compat_string_and_number():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Hello"
        ws["B1"] = 42.5

    rb = _write_and_read(setup)
    ws = rb.active
    assert ws["A1"].value == "Hello"
    assert ws["B1"].value == 42.5


def test_compat_multiple_sheets():
    def setup(wb):
        wb.active.title = "First"
        wb.active["A1"] = "one"
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "two"

    rb = _write_and_read(setup)
    assert rb.sheetnames == ["First", "Second"]
    assert rb["First"]["A1"].value == "one"
    assert rb["Second"]["A1"].value == "two"


def test_compat_bold_font():
    def setup(wb):
        ws = wb.active
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True)

    rb = _write_and_read(setup)
    assert rb.active["A1"].value == "Bold"
    assert rb.active["A1"].font.bold is True


def test_compat_number_format():
    def setup(wb):
        ws = wb.active
        ws["A1"] = 1234.5
        ws["A1"].number_format = "#,##0.00"

    rb = _write_and_read(setup)
    assert rb.active["A1"].value == 1234.5
    assert rb.active["A1"].number_format == "#,##0.00"


def test_compat_boolean():
    def setup(wb):
        ws = wb.active
        ws["A1"] = True
        ws["A2"] = False

    rb = _write_and_read(setup)
    assert rb.active["A1"].value is True
    assert rb.active["A2"].value is False
