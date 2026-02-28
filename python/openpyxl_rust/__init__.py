from openpyxl_rust.cell import Cell
from openpyxl_rust.rich_text import CellRichText, TextBlock
from openpyxl_rust.chart import (
    AreaChart,
    AreaChart3D,
    BarChart,
    BarChart3D,
    DoughnutChart,
    LineChart,
    LineChart3D,
    PieChart,
    PieChart3D,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
    StockChart,
)
from openpyxl_rust.comments import Comment
from openpyxl_rust.datavalidation import DataValidation
from openpyxl_rust.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, DuplicateRule, FormulaRule, IconSetRule, TextRule, Top10Rule
from openpyxl_rust.image import Image
from openpyxl_rust.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl_rust.page_break import Break, BreakList
from openpyxl_rust.properties import DocumentProperties
from openpyxl_rust.protection import SheetProtection
from openpyxl_rust.styles.protection import Protection
from openpyxl_rust.table import Table, TableColumn, TableStyleInfo
from openpyxl_rust.workbook import DefinedName, Workbook
from openpyxl_rust.worksheet import Worksheet


def load_workbook(filename, data_only=True):
    """Load an xlsx file. Returns a Workbook.

    Args:
        filename: A file path (str or Path), or a file-like object with a
                  .read() method (e.g. BytesIO, open file handle).
        data_only: If True (default), uses calamine (Rust) for fast reading
                   (values only, no formatting). If False, uses openpyxl under
                   the hood for full formatting preservation (read-modify-write).
    """
    if not data_only:
        import openpyxl as _openpyxl

        from openpyxl_rust.loader import _convert_openpyxl_to_rust

        src = _openpyxl.load_workbook(filename, data_only=False)
        return _convert_openpyxl_to_rust(src)
    import os
    from datetime import datetime

    from openpyxl_rust._openpyxl_rust import _load_workbook, _load_workbook_bytes

    # File-like object (BytesIO, open file handle, etc.)
    if hasattr(filename, "read"):
        try:
            raw = filename.read()
        except (UnicodeDecodeError, ValueError) as exc:
            raise TypeError("File-like object must be opened in binary mode (read() must return bytes)") from exc
        if not isinstance(raw, bytes):
            raise TypeError("File-like object must be opened in binary mode (read() must return bytes)")
        data = _load_workbook_bytes(raw)
    # Path-like object (str, pathlib.Path, os.PathLike)
    elif isinstance(filename, (str, os.PathLike)):
        data = _load_workbook(str(filename))
    else:
        raise TypeError(
            f"filename must be a file path (str/Path) or a file-like object "
            f"with a .read() method, got {type(filename).__name__}"
        )
    sheet_names = list(data["sheet_names"])
    sheets_data = data["sheets"]

    wb = Workbook()
    wb._sheets = []  # clear default sheet

    for i, name in enumerate(sheet_names):
        if i == 0:
            # Reuse the default sheet (index 0) that RustWorkbook creates
            sheet_idx = 0
            wb._rust_wb.set_sheet_title(0, name)
        else:
            sheet_idx = wb._rust_wb.add_sheet(name)

        ws = Worksheet(title=name, workbook=wb, sheet_idx=sheet_idx)
        rows = sheets_data[name]
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                if value is not None:
                    # Parse datetime strings from calamine back to Python datetime
                    if isinstance(value, str):
                        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                            try:
                                value = datetime.strptime(value, fmt)
                                if fmt == "%Y-%m-%d":
                                    value = value.date()
                                break
                            except ValueError:
                                continue
                    ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
        wb._sheets.append(ws)

    return wb


__all__ = [
    "AreaChart",
    "AreaChart3D",
    "BarChart",
    "BarChart3D",
    "Cell",
    "CellRichText",
    "CellIsRule",
    "ColorScaleRule",
    "Comment",
    "DataBarRule",
    "DataValidation",
    "Break",
    "BreakList",
    "DefinedName",
    "DocumentProperties",
    "DoughnutChart",
    "DuplicateRule",
    "FormulaRule",
    "IconSetRule",
    "TextRule",
    "Top10Rule",
    "Image",
    "LineChart",
    "LineChart3D",
    "PageMargins",
    "PieChart",
    "PieChart3D",
    "PrintOptions",
    "PrintPageSetup",
    "Protection",
    "RadarChart",
    "Reference",
    "ScatterChart",
    "Series",
    "SheetProtection",
    "StockChart",
    "Table",
    "TableColumn",
    "TableStyleInfo",
    "TextBlock",
    "Workbook",
    "Worksheet",
    "load_workbook",
]
